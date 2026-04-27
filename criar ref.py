import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

EXCEL_PATH = r"D:\ProjPREÇOSCONCORRENCIA\Emdesenvolvimento\links_concorrentes.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)
thin = Side(style="thin", color="CCCCCC")
default_border = Border(left=thin, right=thin, top=thin, bottom=thin)

for sheet_name in wb.sheetnames:
    if sheet_name.lower() == "douromed":
        continue
    ws = wb[sheet_name]
    
    if ws.cell(row=1, column=4).value == "Referência":
        print(f"Folha '{sheet_name}' já tem coluna Referência.")
        continue
        
    print(f"Adicionando coluna Referência à folha '{sheet_name}'...")
    
    # Cabeçalho
    url_header = ws.cell(row=1, column=3)
    ref_header = ws.cell(row=1, column=4, value="Referência")
    if url_header.fill:
        ref_header.fill = PatternFill(fill_type=url_header.fill.fill_type, fgColor=url_header.fill.fgColor)
    if url_header.font:
        ref_header.font = Font(color=url_header.font.color, bold=url_header.font.bold, 
                               name=url_header.font.name, size=url_header.font.size)
    if url_header.alignment:
        ref_header.alignment = Alignment(horizontal=url_header.alignment.horizontal, 
                                         vertical=url_header.alignment.vertical)
    ref_header.border = default_border

    # Linhas de dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
        if row[0].value is None:
            continue
        url_cell = row[2]
        ref_cell = ws.cell(row=row[0].row, column=4, value="")
        if url_cell.fill:
            ref_cell.fill = PatternFill(fill_type=url_cell.fill.fill_type, fgColor=url_cell.fill.fgColor)
        if url_cell.font:
            ref_cell.font = Font(color=url_cell.font.color, bold=False, 
                                 name=url_cell.font.name, size=url_cell.font.size)
        if url_cell.alignment:
            ref_cell.alignment = Alignment(horizontal='left', vertical='center')
        ref_cell.border = default_border
        ref_cell.number_format = '@'   # Formato texto

    ws.column_dimensions['D'].width = 20

wb.save(EXCEL_PATH)
print("\n✅ Coluna 'Referência' adicionada a todas as folhas de concorrentes (formato texto).")