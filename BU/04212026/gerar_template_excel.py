"""
gerar_template_excel.py
Cria/actualiza o ficheiro Excel de links com:
  - Folha 1 "Douromed": catálogo actualizado diariamente do SQL
  - Folhas de concorrentes: criadas se não existirem; preserva links já preenchidos

Uso:
    python gerar_template_excel.py          # actualiza tudo
    python gerar_template_excel.py --dm     # só actualiza folha Douromed
"""

import argparse
import os
import sys
import json
import pyodbc
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Forçar encoding UTF-8 para evitar erros com emojis
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ─────────────────────────────────────────────
#  CONFIGURAÇÃO
# ─────────────────────────────────────────────

CAMINHO = r"D:\ProjPREÇOSCONCORRENCIA\Emdesenvolvimento\links_concorrentes.xlsx"
CONFIG_JSON = r"D:\ProjPREÇOSCONCORRENCIA\Emdesenvolvimento\concorrentes_ativos.json"

SQL_SERVER = "192.168.33.4\\PRIV10"
SQL_DB     = "pridouro"
SQL_USER   = "DM"
SQL_PWD    = "DM"

QUERY_DOUROMED = """
    SELECT DISTINCT
        [Artigo].[Artigo],
        [ArtigoIdioma].[Descricao],
        [ArtigoMoeda].[PVP1],
        [Artigo].[STKActual],
        [Artigo].[STKReposicao],
        [Artigo].[Marca],
        [Familias].[Descricao] as Familia,
        [Artigo].[CDU_RefFornecedor]
    FROM [Artigo]
    LEFT JOIN [ArtigoMoeda]
        ON [Artigo].[Artigo] = [ArtigoMoeda].[Artigo]
    LEFT JOIN [ArtigoIdioma]
        ON [Artigo].[Artigo] = [ArtigoIdioma].[Artigo]
        AND [ArtigoIdioma].[Idioma] = 'PT'
    LEFT JOIN [Familias] WITH (NOLOCK)
        ON [Artigo].[Familia] = [Familias].[Familia]
    WHERE
        [Artigo].[ArtigoAnulado] = 'False'
"""

# Lista de concorrentes predefinida (usada apenas se o JSON não existir)
CONCORRENTES_PADRAO = [
    "ES_HenrySchein",
    "ES_DvdDental",
    "ES_DentalExpress",
    "ES_Proclinic",
    "ES_RoyalDent",
    "ES_Dentaltix",
    "PT_HenrySchein",
    "PT_Dentaleader",
    "PT_DentalExpress",
    "PT_Montellano",
    "PT_DentalIberica",
    "PT_Dontalia",
    "PT_BNH",
    "PT_AugustoCabral",
    "PT_Dotamed",
    "PT_Exomed",
    "PT_Minhomedica"
]

COR_DOUROMED   = "0D3B6E"
COR_CONC       = "1E3A5F"
COR_LINHAS_ALT = "F2F2F2"
COR_NOVO       = "E2EFDA"

# ─────────────────────────────────────────────
#  HELPERS DE ESTILO
# ─────────────────────────────────────────────

def _borda():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _cabecalho(ws, row, col, texto, cor):
    c = ws.cell(row=row, column=col, value=texto)
    c.fill = PatternFill("solid", fgColor=cor)
    c.font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _borda()

def _celula(ws, row, col, valor, bg="FFFFFF", bold=False, num_format=None):
    c = ws.cell(row=row, column=col, value=valor)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Calibri", size=10, bold=bold)
    c.alignment = Alignment(vertical="center")
    c.border = _borda()
    if num_format:
        c.number_format = num_format

# ─────────────────────────────────────────────
#  CARREGAR LISTA DE CONCORRENTES ATIVOS
# ─────────────────────────────────────────────

def carregar_concorrentes_ativos():
    """Carrega a lista de concorrentes a partir do ficheiro JSON."""
    if os.path.exists(CONFIG_JSON):
        try:
            with open(CONFIG_JSON, "r", encoding="utf-8") as f:
                dados = json.load(f)
                # As chaves do dicionário são os nomes dos concorrentes
                return list(dados.keys())
        except Exception as e:
            print(f"[AVISO] Erro ao ler {CONFIG_JSON}: {e}. A usar lista padrão.")
            return CONCORRENTES_PADRAO
    else:
        print("[INFO] Ficheiro de configuração não encontrado. A usar lista padrão.")
        return CONCORRENTES_PADRAO

# ─────────────────────────────────────────────
#  SQL → DOUROMED
# ─────────────────────────────────────────────

def carregar_douromed():
    print("[INFO] Ligando ao SQL Server...")
    conn_str = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={SQL_SERVER};"
        f"DATABASE={SQL_DB};"
        f"UID={SQL_USER};"
        f"PWD={SQL_PWD};"
    )
    try:
        conn = pyodbc.connect(conn_str, timeout=15)
        df = pd.read_sql(QUERY_DOUROMED, conn)
        conn.close()
        df.columns = ["Artigo", "Descricao", "PVP1", "STKActual",
                      "STKReposicao", "Marca", "Familia", "RefFornecedor"]
        df["Artigo"] = df["Artigo"].astype(str).str.strip()
        print(f"   [OK] {len(df)} artigos carregados do SQL")
        return df
    except Exception as e:
        print(f"   [ERRO] SQL: {e}")
        return None

# ─────────────────────────────────────────────
#  FOLHA DOUROMED (sempre actualizada)
# ─────────────────────────────────────────────

def actualizar_folha_douromed(wb, df):
    if "Douromed" in wb.sheetnames:
        del wb["Douromed"]

    ws = wb.create_sheet("Douromed", 0)
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22

    cabecalhos = ["Artigo", "Descrição", "PVP1", "STK Actual",
                  "STK Reposição", "Marca", "Família", "Ref. Fornecedor", "Actualizado em"]
    larguras   = [14, 55, 10, 11, 13, 20, 20, 20, 18]

    for col, (cab, larg) in enumerate(zip(cabecalhos, larguras), 1):
        _cabecalho(ws, 1, col, cab, COR_DOUROMED)
        ws.column_dimensions[get_column_letter(col)].width = larg

    data_atual = datetime.now().strftime("%d/%m/%Y %H:%M")

    for r, (_, row) in enumerate(df.iterrows(), 2):
        bg = COR_LINHAS_ALT if r % 2 == 0 else "FFFFFF"
        _celula(ws, r, 1, row["Artigo"],       bg, bold=True)
        _celula(ws, r, 2, row["Descricao"],     bg)
        _celula(ws, r, 3, row["PVP1"],          bg, num_format='#,##0.00 "€"')
        _celula(ws, r, 4, row["STKActual"],      bg)
        _celula(ws, r, 5, row["STKReposicao"],   bg)
        _celula(ws, r, 6, row["Marca"],          bg)
        _celula(ws, r, 7, row["Familia"],        bg)
        _celula(ws, r, 8, row["RefFornecedor"],  bg)
        _celula(ws, r, 9, data_atual,            bg)

    ws.freeze_panes = "A2"
    print(f"   [OK] Folha 'Douromed' actualizada — {len(df)} artigos")

# ─────────────────────────────────────────────
#  FOLHAS DE CONCORRENTES (preserva dados)
# ─────────────────────────────────────────────

def actualizar_folha_concorrente(wb, nome, df_dm):
    cabecalhos = ["Artigo", "Descricao", "URL"]
    larguras   = [14, 55, 70]

    if nome not in wb.sheetnames:
        # Criar do zero
        ws = wb.create_sheet(title=nome)
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 22

        for col, (cab, larg) in enumerate(zip(cabecalhos, larguras), 1):
            _cabecalho(ws, 1, col, cab, COR_CONC)
            ws.column_dimensions[get_column_letter(col)].width = larg

        for r, (_, row) in enumerate(df_dm.iterrows(), 2):
            bg = COR_LINHAS_ALT if r % 2 == 0 else "FFFFFF"
            _celula(ws, r, 1, row["Artigo"],    bg, bold=True)
            _celula(ws, r, 2, row["Descricao"], bg)
            _celula(ws, r, 3, "",               bg)

        ws.freeze_panes = "A2"
        print(f"   [OK] Folha '{nome}' criada — {len(df_dm)} artigos")

    else:
        # Folha já existe — remove obsoletos, adiciona novos
        ws = wb[nome]

        artigos_douromed = set(df_dm["Artigo"].astype(str).str.strip())

        # 1. ELIMINAR artigos que já não existem no Douromed
        linhas_a_apagar = []
        for row in ws.iter_rows(min_row=2, max_col=1):
            val = row[0].value
            if val is not None and str(val).strip() not in artigos_douromed:
                linhas_a_apagar.append(row[0].row)

        for linha in sorted(linhas_a_apagar, reverse=True):
            ws.delete_rows(linha)

        if linhas_a_apagar:
            print(f"   [REMOVER] Folha '{nome}': {len(linhas_a_apagar)} artigos eliminados")

        # 2. ADICIONAR artigos novos
        artigos_existentes = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                artigos_existentes.add(str(row[0]).strip())

        novos = df_dm[~df_dm["Artigo"].isin(artigos_existentes)]

        if novos.empty and not linhas_a_apagar:
            print(f"   [OK] Folha '{nome}': sem alteracoes")
            return

        proxima = ws.max_row + 1
        for r_off, (_, row) in enumerate(novos.iterrows()):
            r = proxima + r_off
            _celula(ws, r, 1, row["Artigo"],    COR_NOVO, bold=True)
            _celula(ws, r, 2, row["Descricao"], COR_NOVO)
            _celula(ws, r, 3, "",               COR_NOVO)

        if novos.shape[0] > 0:
            print(f"   [OK] Folha '{nome}': {len(novos)} artigos novos adicionados (a verde)")

# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dm", action="store_true",
                        help="Actualiza so a folha Douromed")
    args = parser.parse_args()

    print(f"\n{'='*55}")
    print(f"  DENTAL MARKET - Gerar/Actualizar Template Excel")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*55}\n")

    df_sql = carregar_douromed()
    if df_sql is None:
        print("[ERRO] Sem dados SQL. Abortando.")
        return

    if os.path.exists(CAMINHO):
        print("[INFO] Ficheiro existente — a actualizar...")
        wb = openpyxl.load_workbook(CAMINHO)
    else:
        print("[INFO] Ficheiro novo...")
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    print("\n[1/2] Folha Douromed:")
    actualizar_folha_douromed(wb, df_sql)

    if not args.dm:
        print(f"\n[2/2] Folhas de concorrentes:")
        concorrentes_ativos = carregar_concorrentes_ativos()
        print(f"[INFO] {len(concorrentes_ativos)} concorrentes activos carregados.")
        for nome in concorrentes_ativos:
            actualizar_folha_concorrente(wb, nome, df_sql)

    os.makedirs(os.path.dirname(CAMINHO), exist_ok=True)
    wb.save(CAMINHO)

    print(f"\n{'='*55}")
    print(f"  [OK] Guardado: {CAMINHO}")
    print(f"{'='*55}\n")

if __name__ == "__main__":
    main()