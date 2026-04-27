"""
dashboard.py - Dashboard Dental Market
Corre com: streamlit run dashboard.py
"""

import sqlite3
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import plotly.express as px
from datetime import datetime, timedelta
import os
import base64
import subprocess
import sys
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# ============================================================
# CONSTANTES
# ============================================================

DB_PATH = r"D:\ProjPREÇOSCONCORRENCIA\Emdesenvolvimento\historico_precos.db"
EXCEL_PATH = r"D:\ProjPREÇOSCONCORRENCIA\Emdesenvolvimento\links_concorrentes.xlsx"
LOGO_PATH = r"D:\ProjPREÇOSCONCORRENCIA\Emdesenvolvimento\logo.png"

# ============================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================

st.set_page_config(
    page_title="Dental Market",
    page_icon=LOGO_PATH,
    layout="wide"
)

# ============================================================
# CSS GLOBAL
# ============================================================

st.markdown("""
<style>
    html, body, [class*="css"] { font-family: 'Aptos', 'Segoe UI', 'Calibri', sans-serif; }
    .block-container { padding-top: 0.5rem !important; padding-bottom: 0.5rem !important; }
    header[data-testid="stHeader"] { height: 0; }
    div[data-testid="metric-container"] { padding: 4px 8px !important; }
    div[data-testid="stMetricValue"] { font-size: 1.2rem !important; }
    div[data-testid="stMetricLabel"] { font-size: 0.7rem !important; }
    div[data-testid="stVerticalBlock"] > div { gap: 0.2rem !important; }
    .stTabs [data-baseweb="tab-list"] { gap: 4px; margin-bottom: 0 !important; }
    .stTabs [data-baseweb="tab"] { padding: 4px 12px !important; }
    hr { margin: 0.3rem 0 !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"]:has(span[title*="PT |"]) { background-color: #1E3A5F !important; border-color: #1E3A5F !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"]:has(span[title*="PT |"]) span { color: white !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"]:has(span[title*="ES |"]) { background-color: #E67E22 !important; border-color: #E67E22 !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"]:has(span[title*="ES |"]) span { color: white !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"] svg { fill: white !important; }
    .min-price { color: #27ae60 !important; font-weight: 700 !important; }
    .max-price { color: #e74c3c !important; font-weight: 600 !important; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# LOGO
# ============================================================

@st.cache_data
def get_logo_base64():
    if os.path.exists(LOGO_PATH):
        with open(LOGO_PATH, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    return None

logo_base64 = get_logo_base64()

# ============================================================
# FLAGS E NOMES
# ============================================================

def formatar_nome_concorrente(nome: str) -> str:
    if nome is None: return ""
    n = nome.upper()
    if n.startswith("PT_"): return "PT | " + nome[3:]
    elif n.startswith("ES_"): return "ES | " + nome[3:]
    return nome

def flag_html(nome: str) -> str:
    n = nome.upper()
    if n.startswith("PT_"):
        flag = "<img src='https://flagcdn.com/16x12/pt.png' style='vertical-align:middle;margin-right:4px;'>PT"
        label = nome[3:]
    elif n.startswith("ES_"):
        flag = "<img src='https://flagcdn.com/16x12/es.png' style='vertical-align:middle;margin-right:4px;'>ES"
        label = nome[3:]
    else:
        flag = ""
        label = nome
    return f"{flag} {label}" if flag else label

# ============================================================
# LIGAÇÃO À BASE DE DADOS
# ============================================================

@st.cache_resource
def obter_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA cache_size=-32000")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_artigo_conc_data ON precos(artigo, concorrente, data DESC)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sucesso ON precos(sucesso)")
    conn.commit()
    return conn

# ============================================================
# DOUROMED (CATÁLOGO)
# ============================================================

@st.cache_data(ttl=3600)
def carregar_douromed():
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Douromed", dtype=str)
        df.columns = df.columns.str.strip()
        col_map = {c.lower().replace("ã", "a").replace("é", "e").replace("í", "i"): c for c in df.columns}
        col_artigo = col_map.get("artigo", df.columns[0])
        col_descricao = col_map.get("descricao", col_map.get("descrição", df.columns[1] if len(df.columns) > 1 else None))
        col_pvp = next((col_map[k] for k in col_map if "pvp" in k), None)
        col_marca = col_map.get("marca", None)
        col_familia = next((col_map[k] for k in col_map if "famil" in k), None)

        cols_to_keep = [col_artigo]
        if col_descricao: cols_to_keep.append(col_descricao)
        if col_pvp: cols_to_keep.append(col_pvp)
        if col_marca: cols_to_keep.append(col_marca)
        if col_familia: cols_to_keep.append(col_familia)

        df = df[cols_to_keep].copy()

        rename_map = {col_artigo: "artigo"}
        if col_descricao: rename_map[col_descricao] = "descricao"
        if col_pvp: rename_map[col_pvp] = "pvp1"
        if col_marca: rename_map[col_marca] = "marca"
        if col_familia: rename_map[col_familia] = "familia"

        df = df.rename(columns=rename_map)
        df["artigo"] = df["artigo"].astype(str).str.strip()
        if "pvp1" in df.columns: df["pvp1"] = pd.to_numeric(df["pvp1"], errors="coerce")
        if "descricao" in df.columns: df["descricao"] = df["descricao"].fillna("").astype(str)

        return df
    except Exception as e:
        st.warning(f"Não foi possível carregar Douromed: {e}")
        return pd.DataFrame(columns=["artigo", "descricao", "pvp1", "marca", "familia"])

# ============================================================
# LINKS DOS CONCORRENTES
# ============================================================

@st.cache_data(ttl=3600)
def carregar_links():
    try:
        xl = pd.ExcelFile(EXCEL_PATH)
        sheets = [s for s in xl.sheet_names if s.lower() != "douromed"]
        frames = []
        for sheet in sheets:
            df = pd.read_excel(EXCEL_PATH, sheet_name=sheet, dtype=str).fillna("")
            df.columns = df.columns.str.strip()
            col_map = {c.lower(): c for c in df.columns}
            col_art = col_map.get("artigo", df.columns[0])
            col_url = col_map.get("url", df.columns[2] if len(df.columns) > 2 else None)
            if col_url:
                sub = df[[col_art, col_url]].copy()
                sub.columns = ["artigo", "url"]
                sub["concorrente"] = sheet
                sub["artigo"] = sub["artigo"].astype(str).str.strip()
                frames.append(sub[sub["url"].str.startswith("http", na=False)])
        if frames:
            return pd.concat(frames, ignore_index=True)
    except Exception:
        pass
    return pd.DataFrame(columns=["artigo", "url", "concorrente"])

# ============================================================
# QUERIES SQL
# ============================================================

@st.cache_data(ttl=300)
def query_kpis():
    conn = obter_conn()
    return conn.execute("""
        SELECT COUNT(DISTINCT artigo), COUNT(DISTINCT concorrente),
               COUNT(*), MAX(data)
        FROM precos WHERE sucesso=1
    """).fetchone()

@st.cache_data(ttl=300)
def query_concorrentes():
    conn = obter_conn()
    rows = conn.execute("SELECT DISTINCT concorrente FROM precos WHERE sucesso=1 ORDER BY concorrente").fetchall()
    return [r[0] for r in rows]

@st.cache_data(ttl=300)
def query_artigos():
    conn = obter_conn()
    rows = conn.execute("SELECT DISTINCT artigo FROM precos WHERE sucesso=1 ORDER BY artigo").fetchall()
    return [r[0] for r in rows]

@st.cache_data(ttl=300)
def artigos_por_pesquisa(pesquisa: str) -> set:
    if not pesquisa or len(pesquisa) < 2:
        return set()
    conn = obter_conn()
    df = pd.read_sql(
        "SELECT DISTINCT artigo FROM precos WHERE sucesso=1 AND (artigo LIKE ? OR descricao LIKE ?)",
        conn, params=[f"%{pesquisa}%", f"%{pesquisa}%"]
    )
    return set(df["artigo"].astype(str))

@st.cache_data(ttl=300)
def query_comparacao(concorrentes_filtro: tuple, pesquisa: str, marcas: tuple, familias: tuple):
    conn = obter_conn()
    conc_ph = ",".join("?" * len(concorrentes_filtro))
    like = f"%{pesquisa}%" if pesquisa else "%"
    df = pd.read_sql(f"""
        WITH ranked AS (
            SELECT artigo, descricao, concorrente, CAST(preco AS REAL) as preco, url, data,
                   ROW_NUMBER() OVER (PARTITION BY artigo, concorrente ORDER BY data DESC) as rn
            FROM precos WHERE sucesso=1 AND concorrente IN ({conc_ph}) AND (artigo LIKE ? OR descricao LIKE ?)
        )
        SELECT artigo, descricao, concorrente, preco, url, data FROM ranked WHERE rn=1
    """, conn, params=list(concorrentes_filtro) + [like, like])
    df["preco"] = pd.to_numeric(df["preco"], errors="coerce")
    return df

@st.cache_data(ttl=300)
def query_historico(artigo: str, concorrentes_filtro: tuple, dias: int):
    if not concorrentes_filtro:
        return pd.DataFrame()
    conn = obter_conn()
    data_inicio = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")
    placeholders = ",".join("?" * len(concorrentes_filtro))
    df = pd.read_sql(f"""
        SELECT data, concorrente, CAST(preco AS REAL) as preco, promo
        FROM precos WHERE sucesso=1 AND artigo=? AND data>=? AND concorrente IN ({placeholders}) ORDER BY data
    """, conn, params=[artigo, data_inicio] + list(concorrentes_filtro))
    df["data"] = pd.to_datetime(df["data"])
    df["preco"] = pd.to_numeric(df["preco"], errors="coerce")
    return df

@st.cache_data(ttl=300)
def query_alertas():
    conn = obter_conn()
    return pd.read_sql("""
        WITH ranked AS (
            SELECT artigo, descricao, concorrente, CAST(preco AS REAL) as preco, data,
                   ROW_NUMBER() OVER (PARTITION BY artigo, concorrente ORDER BY data DESC) as rn
            FROM precos WHERE sucesso=1
        ),
        atual AS (SELECT * FROM ranked WHERE rn=1),
        ant AS (SELECT * FROM ranked WHERE rn=2)
        SELECT a.artigo, a.descricao, a.concorrente,
               a.preco as preco_atual, p.preco as preco_anterior,
               ROUND((a.preco-p.preco)/p.preco*100,1) as variacao_pct
        FROM atual a JOIN ant p USING(artigo,concorrente)
        WHERE a.preco IS NOT NULL AND p.preco IS NOT NULL AND a.preco<>p.preco ORDER BY variacao_pct
    """, conn)

# ============================================================
# FUNÇÕES DE GESTÃO DE CONCORRENTES
# ============================================================

def eliminar_concorrente(concorrente):
    try:
        conn = obter_conn()
        cursor = conn.execute("DELETE FROM precos WHERE concorrente = ?", (concorrente,))
        registos = cursor.rowcount
        conn.commit()
        if os.path.exists(EXCEL_PATH):
            wb = openpyxl.load_workbook(EXCEL_PATH)
            if concorrente in wb.sheetnames:
                del wb[concorrente]
                wb.save(EXCEL_PATH)
                excel_ok = True
            else:
                excel_ok = False
        else:
            excel_ok = False
        return True, registos, excel_ok
    except Exception as e:
        return False, 0, False, str(e)

def adicionar_concorrente(novo_concorrente):
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        if not os.path.exists(EXCEL_PATH):
            return False, "Excel não encontrado"
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if novo_concorrente in wb.sheetnames:
            return False, "Já existe"

        ws = wb.create_sheet(title=novo_concorrente)
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 22

        thin = Side(style="thin", color="CCCCCC")
        def _borda():
            return Border(left=thin, right=thin, top=thin, bottom=thin)
        def _cabecalho(ws, row, col, texto, cor):
            c = ws.cell(row=row, column=col, value=texto)
            c.fill = PatternFill("solid", fgColor=cor)
            c.font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _borda()
        def _celula(ws, row, col, valor, bg="FFFFFF", bold=False):
            c = ws.cell(row=row, column=col, value=valor)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", size=10, bold=bold)
            c.alignment = Alignment(vertical="center")
            c.border = _borda()

        cabecalhos = ["Artigo", "Descricao", "URL"]
        larguras = [14, 55, 70]
        for col, (cab, larg) in enumerate(zip(cabecalhos, larguras), 1):
            _cabecalho(ws, 1, col, cab, "1E3A5F")
            ws.column_dimensions[get_column_letter(col)].width = larg

        df_dm = carregar_douromed()
        if df_dm is not None:
            for r, (_, row) in enumerate(df_dm.iterrows(), 2):
                bg = "F2F2F2" if r % 2 == 0 else "FFFFFF"
                _celula(ws, r, 1, row["artigo"], bg, bold=True)
                _celula(ws, r, 2, row["descricao"], bg)
                _celula(ws, r, 3, "", bg)

        ws.freeze_panes = "A2"
        wb.save(EXCEL_PATH)
        return True, f"Concorrente '{novo_concorrente}' adicionado"
    except Exception as e:
        return False, f"Erro: {str(e)}"

# ============================================================
# CABEÇALHO E MÉTRICAS
# ============================================================

kpis = query_kpis()

if not kpis or kpis[3] is None:
    st.warning("⚠️ Sem dados ainda. Corre primeiro o scraper.py para popular a base de dados.")
    st.info("💡 Dica: Execute 'python scraper.py --teste 5' para testar com alguns produtos.")
    st.stop()

ultima_data = datetime.strptime(kpis[3][:16], "%Y-%m-%d %H:%M")

col_logo, col_title = st.columns([1, 10])
with col_logo:
    if logo_base64:
        st.markdown(f'<div style="display: flex; justify-content: center; align-items: center;"><img src="data:image/png;base64,{logo_base64}" style="max-width: 40px; max-height: 40px;"></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div style="display: flex; justify-content: center; align-items: center;"><span style="font-size: 32px;">🦷</span></div>', unsafe_allow_html=True)
with col_title:
    st.markdown("<h3 style='margin:0; padding:0;'>Dental Market — Preços Concorrência</h3>", unsafe_allow_html=True)

col1, col2, col3, col4 = st.columns(4)
col1.metric("📦 Artigos Analisados", kpis[0])
col2.metric("🏢 Concorrentes Analisados", kpis[1])
col3.metric("📊 Registos BD", f"{kpis[2]:,}".replace(",", " "))
col4.metric("🕐 Última Atualização", ultima_data.strftime("%d/%m/%Y %H:%M"))
st.divider()

# ============================================================
# DADOS AUXILIARES
# ============================================================

df_dm = carregar_douromed()
df_links = carregar_links()
todos_conc = query_concorrentes()
todos_art = query_artigos()

if not df_dm.empty and todos_art:
    artigos_com_preco = set(str(a) for a in todos_art)
    df_dm_com_preco = df_dm[df_dm["artigo"].astype(str).isin(artigos_com_preco)]
else:
    df_dm_com_preco = df_dm

marcas_disp = sorted(df_dm_com_preco["marca"].dropna().unique()) if not df_dm_com_preco.empty else []
familias_disp = sorted(df_dm_com_preco["familia"].dropna().unique()) if not df_dm_com_preco.empty else []

# ============================================================
# CRIAÇÃO DAS TABS
# ============================================================

tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Comparação actual", "📈 Evolução de preços", "🔔 Alertas", "⚙️ Gestão de Concorrentes", "🔗 Gestão de Links"])

# ============================================================
# TAB 1 – COMPARAÇÃO ACTUAL
# ============================================================
with tab1:
    conc_pt = [c for c in todos_conc if c.upper().startswith("PT_")]
    conc_es = [c for c in todos_conc if c.upper().startswith("ES_")]
    conc_outros = [c for c in todos_conc if not c.upper().startswith(("PT_", "ES_"))]

    with st.expander("🔎 Filtros", expanded=True):
        col_s, col_m, col_f = st.columns([2, 2, 2])
        with col_s:
            pesquisa = st.text_input("Artigo / Descrição:", "")
        with col_m:
            if pesquisa and not df_dm_com_preco.empty:
                arts_pesq = artigos_por_pesquisa(pesquisa)
                dm_pesq = df_dm_com_preco[
                    df_dm_com_preco["artigo"].astype(str).isin(arts_pesq) |
                    df_dm_com_preco["artigo"].str.contains(pesquisa, case=False, na=False) |
                    df_dm_com_preco["familia"].str.contains(pesquisa, case=False, na=False) |
                    df_dm_com_preco["marca"].str.contains(pesquisa, case=False, na=False)
                ]
                marcas_cascade = sorted(dm_pesq["marca"].dropna().unique())
            else:
                marcas_cascade = marcas_disp
            marcas_sel = st.multiselect("Marca:", marcas_cascade, placeholder="Escolha Opção")
        with col_f:
            if pesquisa and not df_dm_com_preco.empty:
                if 'arts_pesq' not in locals():
                    arts_pesq = artigos_por_pesquisa(pesquisa)
                dm_base = df_dm_com_preco[
                    df_dm_com_preco["artigo"].astype(str).isin(arts_pesq) |
                    df_dm_com_preco["artigo"].str.contains(pesquisa, case=False, na=False) |
                    df_dm_com_preco["familia"].str.contains(pesquisa, case=False, na=False) |
                    df_dm_com_preco["marca"].str.contains(pesquisa, case=False, na=False)
                ].copy()
            else:
                dm_base = df_dm_com_preco.copy() if not df_dm_com_preco.empty else pd.DataFrame()
            if marcas_sel and not dm_base.empty:
                dm_base = dm_base[dm_base["marca"].isin(marcas_sel)]
            familias_cascade = sorted(dm_base["familia"].dropna().unique()) if not dm_base.empty else familias_disp
            familias_sel = st.multiselect("Família:", familias_cascade, placeholder="Escolha Opção")

        st.divider()

        col_pais, col_conc = st.columns([1, 4])
        with col_pais:
            st.markdown("**País:**")
            inc_pt = st.checkbox("Portugal", value=True, key="chk_pt")
            inc_es = st.checkbox("Espanha", value=True, key="chk_es")
            if conc_outros:
                inc_out = st.checkbox("Outros", value=True, key="chk_out")
            else:
                inc_out = False

        with col_conc:
            conc_por_pais = []
            if inc_pt:
                conc_por_pais += conc_pt
            if inc_es:
                conc_por_pais += conc_es
            if inc_out:
                conc_por_pais += conc_outros
            conc_sel = st.multiselect("Concorrentes:", conc_por_pais, default=conc_por_pais, format_func=formatar_nome_concorrente)

    if not conc_sel:
        st.info("Selecciona pelo menos um concorrente.")
    else:
        artigos_permitidos = None
        if (marcas_sel or familias_sel) and not df_dm.empty:
            mask = pd.Series([True] * len(df_dm), index=df_dm.index)
            if marcas_sel:
                mask &= df_dm["marca"].isin(marcas_sel)
            if familias_sel:
                mask &= df_dm["familia"].isin(familias_sel)
            artigos_permitidos = set(df_dm[mask]["artigo"].astype(str))

        with st.spinner("A carregar preços..."):
            df_raw = query_comparacao(tuple(conc_sel), pesquisa, tuple(marcas_sel), tuple(familias_sel))

        if artigos_permitidos is not None:
            df_raw = df_raw[df_raw["artigo"].astype(str).isin(artigos_permitidos)]

        if df_raw.empty:
            st.info("Nenhum artigo encontrado.")
        else:
            pivot_preco = df_raw.pivot_table(index=["artigo", "descricao"], columns="concorrente", values="preco", aggfunc="first").reset_index()
            pivot_preco.columns.name = None

            if not df_dm.empty:
                dm_sub = df_dm[["artigo", "pvp1", "marca", "familia"]].copy()
                dm_sub["artigo"] = dm_sub["artigo"].astype(str)
                pivot_preco["artigo"] = pivot_preco["artigo"].astype(str)
                pivot_preco = pivot_preco.merge(dm_sub, on="artigo", how="left")

            price_cols = [c for c in pivot_preco.columns if c in conc_sel]
            all_num_cols = price_cols + (["pvp1"] if "pvp1" in pivot_preco.columns else [])
            if all_num_cols:
                nums_all = pivot_preco[all_num_cols].apply(pd.to_numeric, errors="coerce")
                pivot_preco["_min_val"] = nums_all.min(axis=1)
                pivot_preco["_max_val"] = nums_all.max(axis=1)
                pivot_preco["Mínimo €"] = pivot_preco["_min_val"].round(2)

                def get_todos_mais_baratos(row):
                    mn = row["_min_val"]
                    if pd.isna(mn):
                        return "—"
                    res = []
                    if "pvp1" in row and pd.notna(row["pvp1"]) and abs(float(row["pvp1"]) - mn) < 0.005:
                        res.append("Douromed")
                    for c in price_cols:
                        try:
                            if pd.notna(row[c]) and abs(float(row[c]) - mn) < 0.005:
                                res.append(formatar_nome_concorrente(c))
                        except:
                            pass
                    return " · ".join(res) if res else "—"
                pivot_preco["Mais barato"] = pivot_preco.apply(get_todos_mais_baratos, axis=1)

            rename_cols = {c: formatar_nome_concorrente(c) for c in price_cols}
            rename_html = {c: flag_html(c) for c in price_cols}
            pivot_preco = pivot_preco.rename(columns=rename_cols)
            price_cols_fmt = [rename_cols[c] for c in price_cols]

            pivot_preco = pivot_preco.rename(columns={
                "artigo": "Artigo",
                "descricao": "Descrição",
                "pvp1": "PVP Dm",
                "marca": "Marca",
                "familia": "Família"
            })

            def fmt_preco(x):
                try:
                    if pd.isna(x) or x == "" or x == "—":
                        return "—"
                    return f"{float(x):.2f} €"
                except:
                    return "—"

            for col in price_cols_fmt + ["PVP Dm", "Mínimo €"]:
                if col in pivot_preco.columns:
                    pivot_preco[col] = pivot_preco[col].apply(fmt_preco)

            def extrair_num(s):
                try:
                    if pd.isna(s) or s == "—":
                        return None
                    return float(str(s).replace("€", "").replace(",", ".").strip())
                except:
                    return None

            cols_para_cor = price_cols_fmt + (["PVP Dm"] if "PVP Dm" in pivot_preco.columns else [])
            row_min_vals = {}
            row_max_vals = {}
            for idx2, row2 in pivot_preco.iterrows():
                vals = [extrair_num(row2.get(c)) for c in cols_para_cor]
                vals = [v for v in vals if v is not None]
                row_min_vals[idx2] = min(vals) if vals else None
                row_max_vals[idx2] = max(vals) if vals else None

            url_map = {}
            if not df_links.empty:
                for _, lr in df_links.iterrows():
                    url_map[(str(lr["artigo"]), lr["concorrente"])] = lr["url"]

            fmt_to_orig = {v: k for k, v in rename_cols.items()}
            cols_fixas = ["Artigo", "Descrição", "Marca", "Família", "PVP Dm", "Mínimo €", "Mais barato"]
            cols_final = [c for c in cols_fixas + price_cols_fmt if c in pivot_preco.columns]

            # CSS da tabela com scroll horizontal
            css_table = """
            <style>
            .dm-table-wrap { 
                overflow-x: auto !important; 
                max-height: 560px; 
                overflow-y: auto; 
                border: 1px solid #444; 
                border-radius: 8px; 
                width: 100%;
            }
            .dm-table { 
                border-collapse: collapse; 
                width: 100%; 
                min-width: 800px;
                font-family: 'Aptos', 'Segoe UI', 'Calibri', sans-serif; 
                font-size: 13px; 
                color: #e0e0e0; 
            }
            .dm-table thead th { 
                background: #1E3A5F; 
                color: #fff; 
                padding: 8px 12px; 
                white-space: nowrap; 
                position: sticky; 
                top: 0; 
                z-index: 2; 
            }
            .dm-table thead th.sortable { 
                cursor: pointer; 
                user-select: none; 
            }
            .dm-table thead th.sortable:hover { 
                background: #2a5298; 
            }
            .dm-table thead th .si { 
                margin-left: 4px; 
                font-size: 9px; 
                opacity: 0.6; 
            }
            .dm-table thead th.asc .si::after { 
                content: "▲"; 
                opacity: 1; 
            }
            .dm-table thead th.desc .si::after { 
                content: "▼"; 
                opacity: 1; 
            }
            .dm-table thead th.sortable:not(.asc):not(.desc) .si::after { 
                content: "⇅"; 
            }
            .dm-table thead th.left { 
                text-align: left; 
            }
            .dm-table thead th.center { 
                text-align: center; 
            }
            .dm-table tbody tr:nth-child(even) td { 
                background: #2a2a2a; 
            }
            .dm-table tbody tr:nth-child(odd) td { 
                background: #1e1e1e; 
            }
            .dm-table tbody tr:hover td { 
                background: #2e3d50 !important; 
            }
            .dm-table td { 
                padding: 6px 12px; 
                white-space: nowrap; 
                border-bottom: 1px solid #3a3a3a; 
            }
            .dm-table td.left { 
                text-align: left; 
            }
            .dm-table td.center { 
                text-align: center; 
            }
            a.dm-link { 
                color: #5b9bd5; 
                text-decoration: none; 
                font-weight: 500; 
            }
            a.dm-link:hover { 
                text-decoration: underline; 
            }
            .min-price { 
                color: #27ae60 !important; 
                font-weight: 700 !important; 
            }
            .max-price { 
                color: #e74c3c !important; 
                font-weight: 600 !important; 
            }
            </style>
            """

            rename_html_fmt = {rename_cols[c]: rename_html[c] for c in price_cols}
            SORTABLE = {"Artigo", "Descrição", "Marca", "Família"}

            # Cabeçalho da tabela
            th_rows = "<tr>"
            for ci, c in enumerate(cols_final):
                cls = "left" if c in ["Artigo", "Descrição", "Marca", "Família"] else "center"
                label = rename_html_fmt.get(c, c).replace("🦷 ", "")
                if c in SORTABLE:
                    th_rows += f"<th class='{cls} sortable' onclick='sort({ci},this)'>{label}<span class='si'></span></th>"
                else:
                    th_rows += f"<th class='{cls}'>{label}</th>"
            th_rows += "</tr>"

            # Corpo da tabela
            linhas = []
            for idx2, row in pivot_preco.iterrows():
                artigo_val = str(row.get("Artigo", ""))
                row_min = row_min_vals.get(idx2)
                row_max = row_max_vals.get(idx2)
                tr = "<tr>"
                for c in cols_final:
                    val = row.get(c, "")
                    val_str = "—" if (val is None or pd.isna(val) or str(val).strip() in ("", "nan", "None")) else str(val)
                    cls = "left" if c in ["Artigo", "Descrição", "Marca", "Família"] else "center"
                    extra_class = ""
                    num_val = extrair_num(val_str) if c in cols_para_cor and val_str != "—" else None
                    if num_val is not None and row_min is not None:
                        if abs(num_val - row_min) < 0.005:
                            extra_class = "min-price"
                        elif row_max is not None and abs(num_val - row_max) < 0.005 and abs(row_max - row_min) > 0.005:
                            extra_class = "max-price"

                    if c in price_cols_fmt and val_str != "—":
                        orig = fmt_to_orig.get(c, c)
                        if orig in conc_sel:
                            url = url_map.get((artigo_val, orig), "")
                            if url and url.startswith("http"):
                                if extra_class == "min-price":
                                    cell = f"<a href='{url}' target='_blank' style='color:#27ae60;text-decoration:none;font-weight:700;'>{val_str}</a>"
                                elif extra_class == "max-price":
                                    cell = f"<a href='{url}' target='_blank' style='color:#e74c3c;text-decoration:none;font-weight:600;'>{val_str}</a>"
                                else:
                                    cell = f"<a href='{url}' target='_blank' class='dm-link'>{val_str}</a>"
                            else:
                                cell = val_str
                        else:
                            cell = val_str
                    elif c == "PVP Dm" and val_str != "—":
                        if extra_class == "min-price":
                            cell = f"<span style='color:#27ae60;font-weight:700;'>{val_str}</span>"
                        elif extra_class == "max-price":
                            cell = f"<span style='color:#e74c3c;font-weight:600;'>{val_str}</span>"
                        else:
                            cell = val_str
                    elif c == "Mínimo €" and val_str != "—":
                        cell = f"<span style='color:#27ae60;font-weight:700;'>{val_str}</span>"
                    elif c == "Mais barato" and val_str not in ("", "—", "nan"):
                        baratos = val_str.split(" · ")
                        links_html = []
                        for b in baratos:
                            if b == "Douromed":
                                links_html.append('<span style="color:#27ae60;font-weight:700;">Douromed</span>')
                            else:
                                raw_mb = b.replace("PT | ", "PT_").replace("ES | ", "ES_")
                                if raw_mb in conc_sel:
                                    url_mb = None
                                    for (art, conc), url in url_map.items():
                                        if art == artigo_val and (conc == raw_mb or formatar_nome_concorrente(conc) == b):
                                            url_mb = url
                                            break
                                    if url_mb and url_mb.startswith("http"):
                                        links_html.append(f"<a href='{url_mb}' target='_blank' style='color:#27ae60;text-decoration:none;font-weight:700;'>{b}</a>")
                                    else:
                                        links_html.append(f"<span style='color:#27ae60;font-weight:700;'>{b}</span>")
                                else:
                                    links_html.append(f"<span style='color:#27ae60;font-weight:700;'>{b}</span>")
                        cell = " · ".join(links_html) if links_html else "—"
                        extra_class = "min-price"
                    else:
                        cell = val_str

                    class_attr = f" class='{cls}'"
                    if extra_class == "min-price":
                        tr += f"<td{class_attr} style='color:#27ae60;font-weight:700;'>{cell}</td>"
                    elif extra_class == "max-price":
                        tr += f"<td{class_attr} style='color:#e74c3c;font-weight:600;'>{cell}</td>"
                    else:
                        tr += f"<td{class_attr}>{cell}</td>"
                tr += "</tr>"
                linhas.append(tr)
            rows_html = "".join(linhas)

            js_sort = """
            <script>
            function sort(idx, th) {
                var table = th.closest("table");
                var tbody = table.querySelector("tbody");
                var rows = Array.from(tbody.querySelectorAll("tr"));
                var asc = th.classList.contains("desc") || !th.classList.contains("asc");
                table.querySelectorAll("thead th").forEach(function(h){ h.classList.remove("asc","desc"); });
                th.classList.add(asc ? "asc" : "desc");
                rows.sort(function(a,b){
                    var va = (a.cells[idx] ? a.cells[idx].innerText : "").trim();
                    var vb = (b.cells[idx] ? b.cells[idx].innerText : "").trim();
                    return asc ? va.localeCompare(vb,"pt") : vb.localeCompare(va,"pt");
                });
                rows.forEach(function(r){ tbody.appendChild(r); });
            }
            </script>
            """
            n_rows = len(pivot_preco)
            altura = min(600, 80 + n_rows * 35)
            tabela_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{css_table}{js_sort}</head><body style="margin:0;padding:0;background:transparent;"><div class="dm-table-wrap" style="max-height:{altura}px;"><table class="dm-table" id="tbl"><thead>{th_rows}</thead><tbody>{rows_html}</tbody></table></div></body></html>"""
            components.html(tabela_html, height=altura + 20, scrolling=False)
            st.caption(f"{n_rows} artigos · {ultima_data.strftime('%d/%m/%Y %H:%M')}")

            if st.button("📥 Exportar para Excel"):
                from io import BytesIO
                buf = BytesIO()
                df_export = pivot_preco[cols_final].copy()
                df_export.to_excel(buf, index=False)
                st.download_button("Descarregar", buf.getvalue(), file_name=f"comparacao_{datetime.now().strftime('%Y%m%d')}.xlsx")

# ============================================================
# TAB 2 – EVOLUÇÃO DE PREÇOS
# ============================================================
with tab2:
    if not todos_art:
        st.info("Nenhum artigo encontrado.")
    else:
        col_a, col_p = st.columns([3, 2])
        with col_a:
            artigo_sel = st.selectbox("Artigo:", todos_art)
        with col_p:
            periodo = st.selectbox("Período:", ["30 dias", "90 dias", "180 dias", "Todo o histórico"], index=1)

        dias_map = {"30 dias": 30, "90 dias": 90, "180 dias": 180, "Todo o histórico": 9999}

        with st.spinner("A carregar histórico..."):
            if 'conc_sel' in locals() and conc_sel:
                conc_para_historico = tuple(conc_sel)
            else:
                conc_para_historico = tuple(todos_conc) if todos_conc else ()
            df_hist = query_historico(artigo_sel, conc_para_historico, dias_map[periodo])

        if df_hist.empty:
            st.info("Sem histórico para este artigo.")
        else:
            pvp_dm = None
            if not df_dm.empty:
                row_dm = df_dm[df_dm["artigo"] == artigo_sel]
                if not row_dm.empty:
                    pvp_dm = row_dm["pvp1"].values[0]

            df_hist["concorrente_fmt"] = df_hist["concorrente"].apply(formatar_nome_concorrente)
            fig = px.line(df_hist, x="data", y="preco", color="concorrente_fmt", markers=True,
                          labels={"data": "Data", "preco": "Preço (€)", "concorrente_fmt": "Concorrente"},
                          title=f"Evolução de preços — {artigo_sel}")

            if pvp_dm and not pd.isna(pvp_dm):
                fig.add_hline(y=pvp_dm, line_dash="dash", line_color="green",
                              annotation_text=f"PVP Douromed: {pvp_dm:.2f}€", annotation_position="right")

            fig.update_layout(legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                              yaxis_tickformat=".2f", hovermode="x unified")
            st.plotly_chart(fig, use_container_width=True)

            df_tab = df_hist.copy()
            df_tab["data"] = df_tab["data"].dt.strftime("%d/%m/%Y %H:%M")
            df_tab["preco"] = df_tab["preco"].apply(lambda x: f"{x:.2f} €" if pd.notna(x) else "—")
            df_tab["promo"] = df_tab["promo"].apply(lambda x: "✓" if x else "")
            df_tab["concorrente_fmt"] = df_tab["concorrente"].apply(formatar_nome_concorrente)
            df_tab = df_tab[["data", "concorrente_fmt", "preco", "promo"]].copy()
            df_tab.columns = ["Data", "Concorrente", "Preço", "Promo"]
            st.dataframe(df_tab.sort_values("Data", ascending=False), use_container_width=True, height=300, hide_index=True)

# ============================================================
# TAB 3 – ALERTAS
# ============================================================
with tab3:
    st.subheader("📊 Variações de preço (última vs penúltima recolha)")

    with st.spinner("A carregar variações de preço..."):
        conn = obter_conn()

        df_alertas = pd.read_sql("""
            WITH precos_ordenados AS (
                SELECT 
                    artigo,
                    descricao,
                    concorrente,
                    preco,
                    data,
                    LAG(preco) OVER (PARTITION BY artigo, concorrente ORDER BY data) as preco_anterior,
                    LAG(data) OVER (PARTITION BY artigo, concorrente ORDER BY data) as data_anterior
                FROM precos 
                WHERE sucesso=1
            )
            SELECT 
                artigo,
                descricao,
                concorrente,
                preco as preco_atual,
                preco_anterior,
                data as data_atual,
                data_anterior,
                ROUND((preco - preco_anterior) / preco_anterior * 100, 1) as variacao_pct
            FROM precos_ordenados
            WHERE preco_anterior IS NOT NULL AND preco != preco_anterior
            ORDER BY data DESC
        """, conn)

        if df_alertas.empty:
            st.info("Sem variações de preço registadas.")
        else:
            df_alertas["data_atual"] = pd.to_datetime(df_alertas["data_atual"])
            df_alertas["data_anterior"] = pd.to_datetime(df_alertas["data_anterior"])
            df_alertas["data_atual_str"] = df_alertas["data_atual"].dt.strftime("%Y-%m-%d")
            df_alertas["hora_atual_str"] = df_alertas["data_atual"].dt.strftime("%H:%M")
            df_alertas["data_anterior_str"] = df_alertas["data_anterior"].dt.strftime("%Y-%m-%d")
            df_alertas["hora_anterior_str"] = df_alertas["data_anterior"].dt.strftime("%H:%M")

            df_douromed = carregar_douromed()
            dict_douromed = {}
            if not df_douromed.empty:
                df_douromed["artigo"] = df_douromed["artigo"].astype(str).str.strip()
                for _, row in df_douromed.iterrows():
                    artigo = row["artigo"]
                    dict_douromed[artigo] = {
                        "descricao": row.get("descricao", ""),
                        "marca": row.get("marca", ""),
                        "familia": row.get("familia", "")
                    }

            marcas_lista = []
            familias_lista = []
            for _, row in df_alertas.iterrows():
                artigo = row['artigo']
                info_dm = dict_douromed.get(artigo, {})
                marcas_lista.append(info_dm.get("marca", ""))
                familias_lista.append(info_dm.get("familia", ""))

            df_alertas["marca"] = marcas_lista
            df_alertas["familia"] = familias_lista

            todos_concorrentes = sorted(df_alertas["concorrente"].unique())
            conc_pt = [c for c in todos_concorrentes if c.upper().startswith("PT_")]
            conc_es = [c for c in todos_concorrentes if c.upper().startswith("ES_")]
            conc_outros = [c for c in todos_concorrentes if not c.upper().startswith(("PT_", "ES_"))]

            if "tab3_inc_pt" not in st.session_state:
                st.session_state.tab3_inc_pt = True
            if "tab3_inc_es" not in st.session_state:
                st.session_state.tab3_inc_es = True
            if "tab3_inc_out" not in st.session_state:
                st.session_state.tab3_inc_out = False
            if "tab3_aplicar" not in st.session_state:
                st.session_state.tab3_aplicar = 0
            if "tab3_pesquisa_anterior" not in st.session_state:
                st.session_state.tab3_pesquisa_anterior = ""

            with st.expander("🔎 Filtros", expanded=True):
                col_periodo, col_pesquisa = st.columns([1, 2])
                with col_periodo:
                    periodo = st.selectbox(
                        "📅 Período:",
                        ["7 dias", "15 dias", "30 dias", "60 dias", "90 dias", "Todo o histórico"],
                        index=4,
                        key="tab3_periodo"
                    )

                with col_pesquisa:
                    pesquisa = st.text_input(
                        "🔍 Artigo / Descrição:",
                        key="tab3_pesquisa",
                        placeholder="Digite o código ou descrição..."
                    )
                    if pesquisa != st.session_state.tab3_pesquisa_anterior:
                        st.session_state.tab3_pesquisa_anterior = pesquisa
                        st.session_state.tab3_aplicar += 1

                st.divider()

                col_f2, col_f3 = st.columns(2)

                df_filtrado_temp = df_alertas.copy()
                if periodo != "Todo o histórico":
                    dias_map = {"7 dias": 7, "15 dias": 15, "30 dias": 30, "60 dias": 60, "90 dias": 90}
                    dias = dias_map.get(periodo, 90)
                    data_limite = datetime.now() - timedelta(days=dias)
                    df_filtrado_temp = df_filtrado_temp[df_filtrado_temp["data_atual"] >= data_limite]

                if pesquisa:
                    pesquisa_lower = pesquisa.lower()
                    df_filtrado_temp = df_filtrado_temp[
                        df_filtrado_temp["artigo"].str.lower().str.contains(pesquisa_lower, na=False) |
                        df_filtrado_temp["descricao"].str.lower().str.contains(pesquisa_lower, na=False)
                    ]

                marcas_disponiveis = sorted([m for m in df_filtrado_temp["marca"].dropna().unique() if m and m != "nan"])
                with col_f2:
                    marcas_sel = st.multiselect(
                        "🏷️ Marca:",
                        marcas_disponiveis,
                        key="tab3_marcas",
                        placeholder="Escolha Opção"
                    )

                if marcas_sel:
                    df_filtrado_temp = df_filtrado_temp[df_filtrado_temp["marca"].isin(marcas_sel)]

                familias_disponiveis = sorted([f for f in df_filtrado_temp["familia"].dropna().unique() if f and f != "nan"])
                with col_f3:
                    familias_sel = st.multiselect(
                        "📁 Família:",
                        familias_disponiveis,
                        key="tab3_familias",
                        placeholder="Escolha Opção"
                    )

                if familias_sel:
                    df_filtrado_temp = df_filtrado_temp[df_filtrado_temp["familia"].isin(familias_sel)]

                st.divider()

                col_pais, col_conc = st.columns([1, 4])
                with col_pais:
                    st.markdown("**🌍 País:**")
                    inc_pt = st.checkbox("Portugal", value=st.session_state.tab3_inc_pt, key="tab3_chk_pt")
                    inc_es = st.checkbox("Espanha", value=st.session_state.tab3_inc_es, key="tab3_chk_es")
                    if conc_outros:
                        inc_out = st.checkbox("Outros", value=st.session_state.tab3_inc_out, key="tab3_chk_out")
                    else:
                        inc_out = False

                    pt_mudou = inc_pt != st.session_state.tab3_inc_pt
                    es_mudou = inc_es != st.session_state.tab3_inc_es
                    out_mudou = inc_out != st.session_state.tab3_inc_out

                    st.session_state.tab3_inc_pt = inc_pt
                    st.session_state.tab3_inc_es = inc_es
                    st.session_state.tab3_inc_out = inc_out

                    if pt_mudou or es_mudou or out_mudou:
                        st.session_state.tab3_aplicar += 1

                with col_conc:
                    conc_por_pais = []
                    conc_pt_filtrados = [c for c in conc_pt if c in df_filtrado_temp["concorrente"].unique()]
                    conc_es_filtrados = [c for c in conc_es if c in df_filtrado_temp["concorrente"].unique()]
                    conc_outros_filtrados = [c for c in conc_outros if c in df_filtrado_temp["concorrente"].unique()]

                    if inc_pt:
                        conc_por_pais += conc_pt_filtrados
                    if inc_es:
                        conc_por_pais += conc_es_filtrados
                    if inc_out:
                        conc_por_pais += conc_outros_filtrados

                    multiselect_key = f"tab3_conc_{st.session_state.tab3_aplicar}"
                    default_concs = conc_por_pais.copy()

                    conc_sel = st.multiselect(
                        "🏢 Concorrentes:",
                        conc_por_pais,
                        default=default_concs,
                        format_func=formatar_nome_concorrente,
                        key=multiselect_key,
                        placeholder="Escolha Opção"
                    )

            df_filtrado = df_alertas.copy()
            if periodo != "Todo o histórico":
                dias_map = {"7 dias": 7, "15 dias": 15, "30 dias": 30, "60 dias": 60, "90 dias": 90}
                dias = dias_map.get(periodo, 90)
                data_limite = datetime.now() - timedelta(days=dias)
                df_filtrado = df_filtrado[df_filtrado["data_atual"] >= data_limite]

            if pesquisa:
                pesquisa_lower = pesquisa.lower()
                df_filtrado = df_filtrado[
                    df_filtrado["artigo"].str.lower().str.contains(pesquisa_lower, na=False) |
                    df_filtrado["descricao"].str.lower().str.contains(pesquisa_lower, na=False)
                ]

            if marcas_sel:
                df_filtrado = df_filtrado[df_filtrado["marca"].isin(marcas_sel)]

            if familias_sel:
                df_filtrado = df_filtrado[df_filtrado["familia"].isin(familias_sel)]

            if conc_sel:
                df_filtrado = df_filtrado[df_filtrado["concorrente"].isin(conc_sel)]

            df_filtrado = df_filtrado.sort_values("data_atual", ascending=False)

            if df_filtrado.empty:
                st.info("Nenhuma variação encontrada com os filtros selecionados.")
            else:
                descidas = df_filtrado[df_filtrado["variacao_pct"] < 0].sort_values("data_atual", ascending=False)
                subidas = df_filtrado[df_filtrado["variacao_pct"] > 0].sort_values("data_atual", ascending=False)

                st.caption(f"📊 {len(df_filtrado)} variações encontradas | 📉 {len(descidas)} descidas | 📈 {len(subidas)} subidas")
                st.markdown("---")

                col_d, col_s = st.columns(2)

                with col_d:
                    st.markdown("#### 📉 Descidas de Preço")
                    if descidas.empty:
                        st.info("Sem descidas com os filtros selecionados.")
                    else:
                        for _, r in descidas.iterrows():
                            variacao = r['variacao_pct']
                            cor = "#27ae60"
                            seta = "▼"

                            artigo_display = f"**{r['artigo']}**"
                            if r.get('descricao') and r['descricao'] not in ("", "nan"):
                                artigo_display = f"**{r['artigo']}** - {r['descricao'][:80]}"

                            info_extra = ""
                            if r.get('marca') and r['marca'] not in ("", "nan"):
                                info_extra += f"🏷️ {r['marca']} "
                            if r.get('familia') and r['familia'] not in ("", "nan"):
                                info_extra += f"📁 {r['familia']}"

                            data_hora_atual = f"{r['data_atual_str']} {r['hora_atual_str']}"
                            data_hora_anterior = f"{r['data_anterior_str']} {r['hora_anterior_str']}"

                            st.markdown(
                                f"""{artigo_display}<br>
                                <span style='color:#666; font-size:0.9em;'>{formatar_nome_concorrente(r['concorrente'])}</span><br>
                                <span style='color:#888; font-size:0.8em;'>{info_extra}</span><br>
                                <span style='color:#666; font-size:0.85em;'>📅 {data_hora_anterior} → {data_hora_atual}</span><br>
                                <span style='color:{cor}; font-weight:bold;'>{seta} {r['preco_anterior']:.2f}€ → {r['preco_atual']:.2f}€ ({variacao:+.1f}%)</span>
                                """,
                                unsafe_allow_html=True
                            )
                            st.markdown("---")

                with col_s:
                    st.markdown("#### 📈 Subidas de Preço")
                    if subidas.empty:
                        st.info("Sem subidas com os filtros selecionados.")
                    else:
                        for _, r in subidas.iterrows():
                            variacao = r['variacao_pct']
                            cor = "#e74c3c"
                            seta = "▲"

                            artigo_display = f"**{r['artigo']}**"
                            if r.get('descricao') and r['descricao'] not in ("", "nan"):
                                artigo_display = f"**{r['artigo']}** - {r['descricao'][:80]}"

                            info_extra = ""
                            if r.get('marca') and r['marca'] not in ("", "nan"):
                                info_extra += f"🏷️ {r['marca']} "
                            if r.get('familia') and r['familia'] not in ("", "nan"):
                                info_extra += f"📁 {r['familia']}"

                            data_hora_atual = f"{r['data_atual_str']} {r['hora_atual_str']}"
                            data_hora_anterior = f"{r['data_anterior_str']} {r['hora_anterior_str']}"

                            st.markdown(
                                f"""{artigo_display}<br>
                                <span style='color:#666; font-size:0.9em;'>{formatar_nome_concorrente(r['concorrente'])}</span><br>
                                <span style='color:#888; font-size:0.8em;'>{info_extra}</span><br>
                                <span style='color:#666; font-size:0.85em;'>📅 {data_hora_anterior} → {data_hora_atual}</span><br>
                                <span style='color:{cor}; font-weight:bold;'>{seta} {r['preco_anterior']:.2f}€ → {r['preco_atual']:.2f}€ ({variacao:+.1f}%)</span>
                                """,
                                unsafe_allow_html=True
                            )
                            st.markdown("---")

# ============================================================
# TAB 4 – GESTÃO DE CONCORRENTES
# ============================================================
with tab4:
    st.subheader("⚙️ Gestão de Concorrentes")

    def get_concorrentes_excel():
        try:
            xl = pd.ExcelFile(EXCEL_PATH)
            sheets = [s for s in xl.sheet_names if s.lower() != "douromed"]
            sheets.sort()
            return sheets
        except Exception as e:
            st.error(f"Erro ao ler Excel: {e}")
            return []

    concorrentes_excel = get_concorrentes_excel()

    urls_homepage = {
        "PT_AugustoCabral": "https://www.acabralmd.com/",
        "PT_BNH": "https://www.bnh.pt/",
        "PT_DentalExpress": "https://www.dentalexpress.pt/",
        "PT_DentalIberica": "https://dentaliberica.com/pt/",
        "PT_Dentaleader": "https://www.dentaleader.com/",
        "PT_Dontalia": "https://www.dontalia.pt/",
        "PT_Dotamed": "https://www.dotamedsaojoao.com/",
        "PT_Exomed": "https://www.exomed.pt/",
        "PT_HenrySchein": "https://www.henryschein.pt/",
        "PT_Minhomedica": "https://loja.minhomedica.pt/",
        "PT_Montellano": "https://www.montellano.pt/",
        "PT_NoolDental": "https://www.nooldental.pt/",
        "PT_Nordental": "https://www.nordental.pt/",
        "PT_TacasDental": "https://www.tacasdental.pt/",
        "PT_TropicoFuturo": "https://tropicofuturo.com/",
        "ES_DentalExpress": "https://www.dentalexpress.es/",
        "ES_Dentaltix": "https://www.dentaltix.com/es/",
        "ES_DvdDental": "https://www.dvd-dental.com/",
        "ES_HenrySchein": "https://www.henryschein.es/",
        "ES_Proclinic": "https://www.proclinic.es/",
        "ES_RoyalDent": "https://www.royal-dent.com/",
        "ES_Uppermat": "https://www.uppermat.com/",
    }

    # ============================================================
    # SEÇÃO 1: ATUALIZAR CATÁLOGO
    # ============================================================
    with st.expander("🔄 Actualizar Catálogo (Douromed + Concorrentes)", expanded=False):
        st.markdown("""
        Esta operação irá:
        - 🔄 Actualizar a folha **Douromed** com os dados mais recentes
        - 📝 Actualizar todas as folhas de concorrentes
        """)

        st.info("⚠️ Esta operação pode demorar alguns minutos.")

        if st.button("🔄 Actualizar Catálogo", type="primary", key="btn_actualizar_catalogo"):
            with st.spinner("A actualizar catálogo..."):
                try:
                    script_path = os.path.join(os.path.dirname(EXCEL_PATH), "gerar_template_excel.py")
                    if os.path.exists(script_path):
                        result = subprocess.run(
                            [sys.executable, script_path],
                            capture_output=True,
                            text=True,
                            timeout=300,
                            encoding='utf-8',
                            errors='replace'
                        )

                        if result.returncode == 0:
                            output = result.stdout
                            import re

                            novos_match = re.search(r'Adicionados?:\s*(\d+)', output, re.IGNORECASE)
                            removidos_match = re.search(r'Removidos?:\s*(\d+)', output, re.IGNORECASE)

                            novos = int(novos_match.group(1)) if novos_match else 0
                            removidos = int(removidos_match.group(1)) if removidos_match else 0

                            st.success("✅ Catálogo actualizado com sucesso!")

                            if novos > 0 or removidos > 0:
                                st.markdown("---")
                                st.markdown("**📊 Resumo da actualização:**")

                                col_r1, col_r2 = st.columns(2)
                                with col_r1:
                                    if novos > 0:
                                        st.success(f"➕ {novos} artigos adicionados")
                                    else:
                                        st.info("➕ 0 artigos adicionados")
                                with col_r2:
                                    if removidos > 0:
                                        st.warning(f"➖ {removidos} artigos removidos")
                                    else:
                                        st.info("➖ 0 artigos removidos")
                            else:
                                st.info("📊 Nenhuma alteração detectada")

                            st.caption(f"🕐 {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

                            st.cache_data.clear()
                            time.sleep(2)
                            st.rerun()
                        else:
                            st.error(f"❌ Erro ao actualizar catálogo: {result.stderr}")
                    else:
                        st.error(f"❌ Script não encontrado: {script_path}")
                except Exception as e:
                    st.error(f"❌ Erro: {str(e)}")

    st.divider()

    # ============================================================
    # SEÇÃO 2: ELIMINAR CONCORRENTE
    # ============================================================
    with st.expander("❌ Eliminar Concorrente", expanded=False):
        if concorrentes_excel:
            concorrente_eliminar = st.selectbox(
                "Seleccione o concorrente a eliminar:",
                concorrentes_excel,
                format_func=formatar_nome_concorrente,
                key="eliminar_select"
            )

            if st.button("🗑️ Eliminar", type="primary", key="btn_eliminar_concorrente"):
                if concorrente_eliminar:
                    st.warning(f"Tem certeza que deseja eliminar '{formatar_nome_concorrente(concorrente_eliminar)}'?")
                    confirmar = st.checkbox("Confirmo que quero eliminar este concorrente", key="confirm_eliminar")
                    if confirmar:
                        with st.spinner(f"A eliminar {concorrente_eliminar}..."):
                            sucesso, registos, excel_removido = eliminar_concorrente(concorrente_eliminar)
                            if sucesso:
                                st.success(f"✅ Concorrente eliminado!")
                                st.cache_data.clear()
                                time.sleep(2)
                                st.rerun()
                            else:
                                st.error(f"❌ Erro ao eliminar: {registos}")
        else:
            st.warning("Nenhum concorrente encontrado.")

    st.divider()

    # ============================================================
    # SEÇÃO 3: ADICIONAR NOVO CONCORRENTE
    # ============================================================
    with st.expander("➕ Adicionar Novo Concorrente", expanded=False):
        novo_concorrente = st.text_input(
            "Nome do novo concorrente (ex: PT_NovoConcorrente):",
            placeholder="PT_Nome ou ES_Nome",
            key="novo_concorrente"
        )
        st.caption("O nome deve começar com PT_ ou ES_ seguido do nome do concorrente")

        if st.button("➕ Adicionar", type="primary", key="btn_adicionar_concorrente"):
            if novo_concorrente:
                if not novo_concorrente.startswith(("PT_", "ES_")):
                    st.error("❌ O nome deve começar com PT_ ou ES_")
                else:
                    with st.spinner(f"A adicionar {novo_concorrente}..."):
                        sucesso, mensagem = adicionar_concorrente(novo_concorrente)
                        if sucesso:
                            st.success(f"✅ {mensagem}")
                            st.cache_data.clear()
                            time.sleep(2)
                            st.rerun()
                        else:
                            st.error(f"❌ {mensagem}")

    st.divider()

    # ============================================================
    # SEÇÃO 4: LISTA DE CONCORRENTES
    # ============================================================
    with st.expander("📋 Concorrentes no Ficheiro Excel", expanded=False):
        if concorrentes_excel:
            conc_pt_lista = [c for c in concorrentes_excel if c.upper().startswith("PT_")]
            conc_es_lista = [c for c in concorrentes_excel if c.upper().startswith("ES_")]

            col_pt, col_es = st.columns(2)

            with col_pt:
                st.markdown("**Portugal**")
                if conc_pt_lista:
                    for c in sorted(conc_pt_lista):
                        nome_formatado = formatar_nome_concorrente(c)
                        url = urls_homepage.get(c, "#")
                        if url != "#":
                            st.markdown(f"- [{nome_formatado}]({url})", unsafe_allow_html=True)
                        else:
                            st.markdown(f"- {nome_formatado}")
                else:
                    st.write("*Nenhum*")

            with col_es:
                st.markdown("**Espanha**")
                if conc_es_lista:
                    for c in sorted(conc_es_lista):
                        nome_formatado = formatar_nome_concorrente(c)
                        url = urls_homepage.get(c, "#")
                        if url != "#":
                            st.markdown(f"- [{nome_formatado}]({url})", unsafe_allow_html=True)
                        else:
                            st.markdown(f"- {nome_formatado}")
                else:
                    st.write("*Nenhum*")

            st.caption(f"Total: {len(concorrentes_excel)} concorrentes")

            if st.button("🔄 Recarregar Lista", key="recarregar_lista"):
                st.cache_data.clear()
                st.rerun()
        else:
            st.info("Nenhum concorrente encontrado no ficheiro Excel.")

# ============================================================
# TAB 5 – GESTÃO DE LINKS
# ============================================================
with tab5:
    st.subheader("🔗 Gestão de Links por Artigo")
    st.caption("Adicione ou edite os links dos artigos para cada concorrente")

    df_dm_gestao = carregar_douromed()
    concorrentes_excel = get_concorrentes_excel()

    col_artigo = "artigo"
    col_descricao = "descricao" if "descricao" in df_dm_gestao.columns else None
    col_marca = "marca" if "marca" in df_dm_gestao.columns else None
    col_familia = "familia" if "familia" in df_dm_gestao.columns else None

    if df_dm_gestao.empty:
        st.warning("Não foi possível carregar os artigos da Douromed.")
    elif not concorrentes_excel:
        st.warning("Nenhum concorrente encontrado no ficheiro Excel.")
    else:
        if col_descricao:
            df_dm_gestao[col_descricao] = df_dm_gestao[col_descricao].fillna("").astype(str)

        dict_info_dm = {}
        if not df_dm_gestao.empty:
            for _, row in df_dm_gestao.iterrows():
                dict_info_dm[row["artigo"]] = {
                    "descricao": row.get("descricao", ""),
                    "marca": row.get("marca", ""),
                    "familia": row.get("familia", "")
                }

        artigos_lista = sorted(dict_info_dm.keys())

        @st.cache_data(ttl=3600)
        def carregar_links_por_concorrente(concorrente: str):
            try:
                df = pd.read_excel(EXCEL_PATH, sheet_name=concorrente, dtype=str).fillna("")
                df.columns = df.columns.str.strip()
                if "Artigo" in df.columns and "URL" in df.columns:
                    links_dict = {}
                    for _, row in df.iterrows():
                        artigo = str(row["Artigo"]).strip()
                        url = row["URL"] if pd.notna(row["URL"]) else ""
                        if url in ("nan", "NaN", "None", ""):
                            url = ""
                        links_dict[artigo] = url
                    return links_dict, len(df)
                return {}, 0
            except Exception:
                return {}, 0

        @st.cache_data(ttl=3600)
        def get_artigos_sem_link_count(concorrente: str):
            try:
                df = pd.read_excel(EXCEL_PATH, sheet_name=concorrente, dtype=str).fillna("")
                df.columns = df.columns.str.strip()
                if "Artigo" in df.columns and "URL" in df.columns:
                    total = len(df)
                    sem_url = df[(df["URL"].isna()) | (df["URL"].str.strip() == "") | (df["URL"].str.strip() == "nan")]
                    return len(sem_url), total
                return 0, 0
            except Exception:
                return 0, 0

        def obter_ultimo_preco(artigo: str, concorrente: str) -> tuple:
            try:
                conn = sqlite3.connect(DB_PATH, timeout=10)
                conn.execute("PRAGMA journal_mode=WAL")
                result = conn.execute("""
                    SELECT preco, data FROM precos 
                    WHERE sucesso=1 AND artigo=? AND concorrente=?
                    ORDER BY data DESC LIMIT 1
                """, (artigo, concorrente)).fetchone()
                conn.close()
                if result and result[0] and float(result[0]) > 0:
                    preco_val = float(result[0])
                    data_val = result[1][:10] if result[1] else ""
                    return (f"{preco_val:.2f} € ({data_val})", preco_val, data_val)
                return ("Sem registo de preço", None, None)
            except Exception:
                return ("Erro", None, None)

        if 'conc_sel_gestao' not in st.session_state:
            st.session_state.conc_sel_gestao = None
        if 'artigo_sel_gestao' not in st.session_state:
            st.session_state.artigo_sel_gestao = None
        if 'refresh_key' not in st.session_state:
            st.session_state.refresh_key = 0
        if 'expandido_sem_link' not in st.session_state:
            st.session_state.expandido_sem_link = False

        def on_concorrente_change():
            st.session_state.conc_sel_gestao = st.session_state.conc_widget
            st.session_state.artigo_sel_gestao = None
            st.session_state.refresh_key += 1

        def on_artigo_change():
            st.session_state.artigo_sel_gestao = st.session_state.art_widget
            st.session_state.refresh_key += 1

        with st.expander("🔎 Filtros", expanded=True):
            col1, col2 = st.columns([1, 2])
            with col1:
                opcoes = [None] + concorrentes_excel
                idx_conc = 0
                if st.session_state.conc_sel_gestao in opcoes:
                    idx_conc = opcoes.index(st.session_state.conc_sel_gestao)
                st.selectbox("🏢 Concorrente", opcoes, format_func=lambda x: formatar_nome_concorrente(x) if x else "Selecione...", index=idx_conc, key="conc_widget", on_change=on_concorrente_change)
            with col2:
                st.text_input("🔍 Pesquisar artigo", placeholder="Código ou descrição...", key="pesquisa_input")

        if not st.session_state.conc_sel_gestao:
            st.info("🏢 Selecione um concorrente para começar.")
        else:
            arts_filtrados = artigos_lista
            pesquisa = st.session_state.get("pesquisa_input", "")
            if pesquisa:
                pesquisa_lower = pesquisa.lower()
                arts_filtrados = [a for a in artigos_lista if pesquisa_lower in a.lower() or pesquisa_lower in dict_info_dm.get(a, {}).get("descricao", "").lower()]

            if not arts_filtrados:
                st.info("Nenhum artigo encontrado.")
            else:
                if st.session_state.artigo_sel_gestao not in arts_filtrados:
                    st.session_state.artigo_sel_gestao = arts_filtrados[0]

                idx_art = arts_filtrados.index(st.session_state.artigo_sel_gestao)
                opcoes_formatadas = {a: f"{a} - {dict_info_dm.get(a, {}).get('descricao', '')[:60]}" for a in arts_filtrados}

                artigo_sel = st.selectbox("📦 Artigo", arts_filtrados, format_func=lambda x: opcoes_formatadas.get(x, x), index=idx_art, key="art_widget", on_change=on_artigo_change)

                try:
                    df_conc = pd.read_excel(EXCEL_PATH, sheet_name=st.session_state.conc_sel_gestao, dtype=str)
                    df_conc.columns = df_conc.columns.str.strip()
                    col_art = df_conc.columns[0]
                    col_url = df_conc.columns[2] if len(df_conc.columns) > 2 else None
                    if col_url:
                        mask = df_conc[col_art].astype(str).str.strip() == artigo_sel
                        if mask.any():
                            url_val = df_conc.loc[mask, col_url].iloc[0]
                            url_atual = str(url_val) if pd.notna(url_val) and str(url_val).lower() not in ("", "nan", "none") else ""
                        else:
                            url_atual = ""
                    else:
                        url_atual = ""
                except Exception:
                    url_atual = ""

                info_dm = dict_info_dm.get(artigo_sel, {})
                ultimo_preco, preco_valor, preco_data = obter_ultimo_preco(artigo_sel, st.session_state.conc_sel_gestao)

                st.markdown("---")
                col_info, col_edit = st.columns([1, 2])

                with col_info:
                    st.markdown(f"**📝 Descrição:** {info_dm.get('descricao', '')}")
                    st.markdown(f"**💰 Último Preço:** {ultimo_preco}")
                    if info_dm.get("marca"):
                        st.markdown(f"**🏷️ Marca:** {info_dm['marca']}")
                    if info_dm.get("familia"):
                        st.markdown(f"**📁 Família:** {info_dm['familia']}")
                    st.markdown(f"**🏢 Concorrente:** {formatar_nome_concorrente(st.session_state.conc_sel_gestao)}")

                with col_edit:
                    input_key = f"url_input_{st.session_state.conc_sel_gestao}_{artigo_sel}_{st.session_state.refresh_key}"
                    novo_url = st.text_input("🔗 URL do produto", value=url_atual, placeholder="https://www.exemplo.pt/produto/123", key=input_key)

                    col_btn1, col_btn2, col_btn3 = st.columns(3)
                    with col_btn1:
                        if st.button("💾 Guardar", use_container_width=True):
                            if not novo_url or not novo_url.startswith("http"):
                                st.error("❌ URL inválido")
                            else:
                                try:
                                    wb = openpyxl.load_workbook(EXCEL_PATH)
                                    ws = wb[st.session_state.conc_sel_gestao]
                                    linha = None
                                    for row in ws.iter_rows(min_row=2):
                                        if row[0].value and str(row[0].value).strip() == artigo_sel:
                                            linha = row[0].row
                                            break
                                    if linha:
                                        ws.cell(row=linha, column=3, value=novo_url)
                                        wb.save(EXCEL_PATH)
                                        st.success("✅ Link guardado!")
                                        st.cache_data.clear()
                                        time.sleep(0.3)
                                        st.rerun()
                                    else:
                                        st.error("❌ Artigo não encontrado")
                                except Exception as e:
                                    st.error(f"❌ Erro: {e}")
                    with col_btn2:
                        if st.button("🗑️ Limpar", use_container_width=True):
                            try:
                                wb = openpyxl.load_workbook(EXCEL_PATH)
                                ws = wb[st.session_state.conc_sel_gestao]
                                linha = None
                                for row in ws.iter_rows(min_row=2):
                                    if row[0].value and str(row[0].value).strip() == artigo_sel:
                                        linha = row[0].row
                                        break
                                if linha:
                                    ws.cell(row=linha, column=3, value="")
                                    wb.save(EXCEL_PATH)
                                    st.success("✅ URL removido!")
                                    st.cache_data.clear()
                                    time.sleep(0.3)
                                    st.rerun()
                                else:
                                    st.error("❌ Artigo não encontrado")
                            except Exception as e:
                                st.error(f"❌ Erro: {e}")
                    with col_btn3:
                        if url_atual and url_atual.startswith("http"):
                            st.markdown(f"[🔗 Abrir]({url_atual})", unsafe_allow_html=True)

    st.divider()

    with st.expander("📋 Artigos sem Link", expanded=False):
        if st.session_state.get("conc_sel_gestao"):
            if st.session_state.get("expandido_sem_link", False):
                with st.spinner("A carregar..."):
                    try:
                        df_conc = pd.read_excel(EXCEL_PATH, sheet_name=st.session_state.conc_sel_gestao, dtype=str)
                        df_conc.columns = df_conc.columns.str.strip()
                        col_art = df_conc.columns[0]
                        col_desc = df_conc.columns[1] if len(df_conc.columns) > 1 else None
                        col_url = df_conc.columns[2] if len(df_conc.columns) > 2 else None
                        if col_url:
                            artigos_sem_link = []
                            for _, row in df_conc.iterrows():
                                artigo = str(row[col_art]).strip()
                                url = row[col_url] if pd.notna(row[col_url]) else ""
                                if not url or url.lower() in ("", "nan", "none"):
                                    descricao_art = ""
                                    if col_desc and pd.notna(row[col_desc]):
                                        descricao_art = str(row[col_desc])
                                    artigos_sem_link.append({"Artigo": artigo, "Descrição": descricao_art[:80] if descricao_art else ""})
                            if artigos_sem_link:
                                st.warning(f"📌 {len(artigos_sem_link)} artigos sem link")
                                df_sem_link = pd.DataFrame(artigos_sem_link)
                                st.dataframe(df_sem_link, use_container_width=True, height=400, hide_index=True)
                                from io import BytesIO
                                output = BytesIO()
                                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                    df_sem_link.to_excel(writer, index=False, sheet_name="Artigos sem Link")
                                st.download_button("📥 Exportar Excel", output.getvalue(), file_name=f"artigos_sem_link_{st.session_state.conc_sel_gestao}.xlsx")
                            else:
                                st.success("✅ Todos os artigos têm link!")
                    except Exception as e:
                        st.error(f"Erro: {e}")
            else:
                st.info("Clique para carregar a lista.")
                if st.button("📋 Carregar artigos sem link"):
                    st.session_state.expandido_sem_link = True
                    st.rerun()
        else:
            st.info("Selecione um concorrente.")

# ============================================================
# FOOTER
# ============================================================

st.markdown("---")
st.markdown(
    f"<div style='text-align: center; color: gray;'>"
    f"Dashboard de Preços Concorrência | Actualizado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    f"</div>",
    unsafe_allow_html=True
)