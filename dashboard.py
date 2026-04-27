"""
dashboard.py - Dashboard Dental Market
Corre com: streamlit run dashboard.py
"""

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import plotly.express as px
from datetime import datetime, timedelta
import os
import base64
import subprocess
import sys
import time
import re
import json
from pathlib import Path
from sqlalchemy import create_engine, text
from dotenv import load_dotenv

# ============================================================
# CARREGAR VARIÁVEIS DE AMBIENTE
# ============================================================
load_dotenv()

# ============================================================
# CONSTANTES
# ============================================================
ADMIN_PASSWORD = "1279"  # altere para uma palavra‑passe segura

# Caminhos para ficheiros locais (apenas logs e scripts externos)
BASE_DIR = Path(__file__).parent if "__file__" in dir() else Path(".")
SCRAPER_LOG = BASE_DIR / "scraper_execucoes.log"

# ============================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================
st.set_page_config(
    page_title="Dental Market",
    page_icon="🦷",
    layout="wide"
)

# ============================================================
# CSS GLOBAL
# ============================================================
st.markdown("""
<style>
    html, body, [class*="css"] { font-family: 'Aptos', 'Segoe UI', 'Calibri', sans-serif; }
    .block-container { padding-top: 0.5rem !important; padding-bottom: 0.5rem !important; }

    header[data-testid="stHeader"] {
        background: transparent !important;
        height: 28px !important;
        min-height: 0 !important;
        padding: 0 !important;
        border: none !important;
        box-shadow: none !important;
        pointer-events: none;
    }
    header[data-testid="stHeader"] [data-testid="stStatusWidget"] {
        pointer-events: auto;
        display: flex !important;
        align-items: center;
        justify-content: flex-end;
        padding-right: 1rem;
    }
    [data-testid="stDeployButton"],
    [data-testid="stMainMenu"],
    [data-testid="stToolbarActions"] {
        display: none !important;
    }
    div[data-testid="metric-container"] { padding: 4px 8px !important; }
    div[data-testid="stMetricValue"] { font-size: 1.2rem !important; }
    div[data-testid="stMetricLabel"] { font-size: 0.7rem !important; }
    div[data-testid="stVerticalBlock"] > div { gap: 0.2rem !important; }
    .stTabs [data-baseweb="tab-list"] { gap: 4px; margin-bottom: 0 !important; }
    .stTabs [data-baseweb="tab"] { padding: 4px 12px !important; }
    hr { margin: 0.3rem 0 !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"]:has(span[title*="PT |"]) { background-color: #1E3A5F !important; border-color: #1E3A5F !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"]:has(span[title*="PT |"]) span { color: white !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"]:has(span[title*="ES |"]) { background-color: #E67E22 !important; border-color: #E67E22 !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"]:has(span[title*="ES |"]) span { color: white !important; }
    div[data-testid="stMultiSelect"] span[data-baseweb="tag"] svg { fill: white !important; }
    .min-price { color: #27ae60 !important; font-weight: 700 !important; }
    .max-price { color: #e74c3c !important; font-weight: 600 !important; }
    footer { display: none !important; }
    .main > div { padding-bottom: 0 !important; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# FLAGS E NOMES
# ============================================================
def formatar_nome_concorrente(nome: str) -> str:
    if nome is None: return ""
    n = nome.upper()
    if n.startswith("PT_"): return "PT | " + nome[3:]
    elif n.startswith("ES_"): return "ES | " + nome[3:]
    return nome

def flag_html(nome: str) -> str:
    n = nome.upper()
    if n.startswith("PT_"):
        flag = "<img src='https://flagcdn.com/16x12/pt.png' style='vertical-align:middle;margin-right:4px;'>PT"
        label = nome[3:]
    elif n.startswith("ES_"):
        flag = "<img src='https://flagcdn.com/16x12/es.png' style='vertical-align:middle;margin-right:4px;'>ES"
        label = nome[3:]
    else:
        flag = ""
        label = nome
    return f"{flag} {label}" if flag else label

# ============================================================
# LIGAÇÃO À BASE DE DADOS (PostgreSQL via Neon)
# ============================================================
@st.cache_resource
def obter_conn():
    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        st.error("❌ Variável DATABASE_URL não configurada no ficheiro .env")
        st.stop()
    engine = create_engine(database_url, pool_size=5, max_overflow=10)
    with engine.connect() as conn:
        conn.execute(text("SELECT 1"))
    return engine

# ============================================================
# DOUROMED (CATÁLOGO)
# ============================================================
@st.cache_data(ttl=3600)
def carregar_douromed():
    engine = obter_conn()
    with engine.connect() as conn:
        result = conn.execute(text("SELECT artigo, descricao, pvp1, marca, familia FROM artigos ORDER BY artigo"))
        df = pd.DataFrame(result.fetchall(), columns=result.keys())
    return df if not df.empty else pd.DataFrame(columns=["artigo", "descricao", "pvp1", "marca", "familia"])

# ============================================================
# LINKS DOS CONCORRENTES
# ============================================================
@st.cache_data(ttl=3600)
def carregar_links():
    engine = obter_conn()
    with engine.connect() as conn:
        result = conn.execute(text("SELECT artigo, concorrente, url FROM links WHERE url IS NOT NULL AND url != ''"))
        df = pd.DataFrame(result.fetchall(), columns=result.keys())
    return df if not df.empty else pd.DataFrame(columns=["artigo", "url", "concorrente"])

# ============================================================
# QUERIES SQL
# ============================================================
@st.cache_data(ttl=300)
def query_kpis():
    engine = obter_conn()
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT COUNT(DISTINCT artigo), COUNT(DISTINCT concorrente),
                   COUNT(*), MAX(data)
            FROM precos WHERE sucesso=1
        """))
        row = result.fetchone()
    return row if row else (0, 0, 0, None)

@st.cache_data(ttl=300)
def query_concorrentes():
    engine = obter_conn()
    with engine.connect() as conn:
        result = conn.execute(text("SELECT DISTINCT concorrente FROM precos WHERE sucesso=1 ORDER BY concorrente"))
        rows = result.fetchall()
    return [r[0] for r in rows]

@st.cache_data(ttl=300)
def query_artigos():
    engine = obter_conn()
    with engine.connect() as conn:
        result = conn.execute(text("SELECT DISTINCT artigo FROM precos WHERE sucesso=1 ORDER BY artigo"))
        rows = result.fetchall()
    return [r[0] for r in rows]

@st.cache_data(ttl=300)
def artigos_por_pesquisa(pesquisa: str) -> set:
    if not pesquisa or len(pesquisa) < 2:
        return set()
    engine = obter_conn()
    with engine.connect() as conn:
        result = conn.execute(
            text("SELECT DISTINCT artigo FROM precos WHERE sucesso=1 AND (artigo LIKE :p OR descricao LIKE :p)"),
            {"p": f"%{pesquisa}%"}
        )
        rows = result.fetchall()
    return set(str(r[0]) for r in rows)

@st.cache_data(ttl=300)
def query_comparacao(concorrentes_filtro: tuple, pesquisa: str, marcas: tuple, familias: tuple):
    engine = obter_conn()
    conc_params = {f"c{i}": conc for i, conc in enumerate(concorrentes_filtro)}
    conc_placeholders = ", ".join(f":c{i}" for i in range(len(concorrentes_filtro)))
    like = f"%{pesquisa}%" if pesquisa else "%"
    sql = f"""
        WITH ranked AS (
            SELECT artigo, descricao, concorrente, preco, url, data, referencia,
                   ROW_NUMBER() OVER (PARTITION BY artigo, concorrente ORDER BY data DESC) as rn
            FROM precos
            WHERE sucesso=1
              AND concorrente IN ({conc_placeholders})
              AND (artigo LIKE :like OR descricao LIKE :like)
        )
        SELECT artigo, descricao, concorrente, preco, url, data, referencia
        FROM ranked
        WHERE rn=1
    """
    params = {**conc_params, "like": like}
    with engine.connect() as conn:
        result = conn.execute(text(sql), params)
        df = pd.DataFrame(result.fetchall(), columns=result.keys())
    df["preco"] = pd.to_numeric(df["preco"], errors="coerce")
    return df

@st.cache_data(ttl=300)
def query_historico(artigo: str, concorrentes_filtro: tuple, dias: int):
    if not concorrentes_filtro:
        return pd.DataFrame()
    engine = obter_conn()
    data_inicio = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")
    conc_params = {f"c{i}": conc for i, conc in enumerate(concorrentes_filtro)}
    conc_placeholders = ", ".join(f":c{i}" for i in range(len(concorrentes_filtro)))
    sql = f"""
        SELECT data, concorrente, preco, promo, referencia
        FROM precos
        WHERE sucesso=1
          AND artigo = :artigo
          AND data >= :data_inicio
          AND concorrente IN ({conc_placeholders})
        ORDER BY data
    """
    params = {"artigo": artigo, "data_inicio": data_inicio, **conc_params}
    with engine.connect() as conn:
        result = conn.execute(text(sql), params)
        df = pd.DataFrame(result.fetchall(), columns=result.keys())
    if not df.empty:
        df["data"] = pd.to_datetime(df["data"])
        df["preco"] = pd.to_numeric(df["preco"], errors="coerce")
    return df

@st.cache_data(ttl=300)
def query_alertas():
    engine = obter_conn()
    sql = text("""
        WITH precos_ordenados AS (
            SELECT
                artigo,
                descricao,
                concorrente,
                preco,
                data,
                referencia,
                LAG(preco) OVER (PARTITION BY artigo, concorrente ORDER BY data) AS preco_anterior,
                LAG(data) OVER (PARTITION BY artigo, concorrente ORDER BY data) AS data_anterior
            FROM precos
            WHERE sucesso = 1
        )
        SELECT
            artigo,
            descricao,
            concorrente,
            preco AS preco_atual,
            preco_anterior,
            data AS data_atual,
            data_anterior,
            ROUND((preco::numeric - preco_anterior::numeric) / preco_anterior::numeric * 100, 1) AS variacao_pct
        FROM precos_ordenados
        WHERE preco_anterior IS NOT NULL AND preco <> preco_anterior
        ORDER BY data DESC
    """)
    with engine.connect() as conn:
        result = conn.execute(sql)
        df = pd.DataFrame(result.fetchall(), columns=result.keys())
    if not df.empty:
        df["data_atual"] = pd.to_datetime(df["data_atual"])
        df["data_anterior"] = pd.to_datetime(df["data_anterior"])
    return df

# ============================================================
# FUNÇÕES DE GESTÃO DE CONCORRENTES (BD)
# ============================================================
def eliminar_concorrente_bd(concorrente: str):
    engine = obter_conn()
    try:
        with engine.connect() as conn:
            trans = conn.begin()
            conn.execute(text("DELETE FROM links WHERE concorrente = :c"), {"c": concorrente})
            conn.execute(text("DELETE FROM precos WHERE concorrente = :c"), {"c": concorrente})
            result = conn.execute(text("DELETE FROM concorrentes WHERE nome = :c"), {"c": concorrente})
            trans.commit()
            return True, result.rowcount, True, ""
    except Exception as e:
        return False, 0, False, str(e)

def adicionar_concorrente_bd(nome: str, homepage: str):
    engine = obter_conn()
    try:
        with engine.connect() as conn:
            conn.execute(
                text("INSERT INTO concorrentes (nome, homepage, ativo) VALUES (:nome, :hp, 1) ON CONFLICT (nome) DO NOTHING"),
                {"nome": nome, "hp": homepage}
            )
            conn.commit()
            return True, f"Concorrente '{nome}' adicionado."
    except Exception as e:
        return False, str(e)

def listar_concorrentes_ativos():
    engine = obter_conn()
    with engine.connect() as conn:
        result = conn.execute(text("SELECT nome, homepage FROM concorrentes WHERE ativo = 1 ORDER BY nome"))
        return [(row[0], row[1]) for row in result.fetchall()]

# ============================================================
# CABEÇALHO E MÉTRICAS
# ============================================================
kpis = query_kpis()

if not kpis or kpis[3] is None:
    st.warning("⚠️ Sem dados ainda. Corre primeiro o scraper.py para popular a base de dados.")
    st.info("💡 Dica: Execute 'python scraper.py --teste 5' para testar com alguns produtos.")
    kpis = (0, 0, 0, None)
    ultima_data = "N/A"
else:
    ultima_data = datetime.strptime(kpis[3][:16], "%Y-%m-%d %H:%M")

col_logo, col_title = st.columns([1, 10])
with col_logo:
    st.markdown('<div style="display: flex; justify-content: center; align-items: center;"><span style="font-size: 32px;">🦷</span></div>', unsafe_allow_html=True)
with col_title:
    st.markdown("<h3 style='margin:0; padding:0;'>Dental Market — Preços Concorrência</h3>", unsafe_allow_html=True)

col1, col2, col3, col4 = st.columns(4)
col1.metric("📦 Artigos Analisados", kpis[0])
col2.metric("🏢 Concorrentes Analisados", kpis[1])
col3.metric("📊 Registos BD", f"{kpis[2]:,}".replace(",", " "))
col4.metric("🕐 Última Atualização", ultima_data.strftime("%d/%m/%Y %H:%M") if ultima_data != "N/A" else "N/A")
st.divider()

col_btn_refresh, col_space = st.columns([1, 11])
with col_btn_refresh:
    if st.button("🔄 Atualizar", help="Limpa o cache e recarrega todos os dados"):
        st.cache_data.clear()
        st.rerun()

# ============================================================
# DADOS AUXILIARES
# ============================================================
df_dm = carregar_douromed()
df_links = carregar_links()
todos_conc = query_concorrentes()
todos_art = query_artigos()

if not df_dm.empty and todos_art:
    artigos_com_preco = set(str(a) for a in todos_art)
    df_dm_com_preco = df_dm[df_dm["artigo"].astype(str).isin(artigos_com_preco)]
else:
    df_dm_com_preco = df_dm

marcas_disp = sorted(df_dm_com_preco["marca"].dropna().unique()) if not df_dm_com_preco.empty else []
familias_disp = sorted(df_dm_com_preco["familia"].dropna().unique()) if not df_dm_com_preco.empty else []

# ============================================================
# CRIAÇÃO DAS TABS
# ============================================================
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📊 Comparação actual",
    "📈 Evolução de preços",
    "🔔 Alertas",
    "⚙️ Gestão de Concorrentes",
    "🔗 Gestão de Links",
    "🛠️ Administração"
])

# ============================================================
# TAB 1 – COMPARAÇÃO ACTUAL
# ============================================================
with tab1:
    conc_pt = [c for c in todos_conc if c.upper().startswith("PT_")]
    conc_es = [c for c in todos_conc if c.upper().startswith("ES_")]
    conc_outros = [c for c in todos_conc if not c.upper().startswith(("PT_", "ES_"))]

    with st.expander("🔎 Filtros", expanded=True):
        col_s, col_m, col_f = st.columns([2, 2, 2])
        with col_s:
            pesquisa = st.text_input("Artigo / Descrição:", "")
        with col_m:
            if pesquisa and not df_dm_com_preco.empty:
                arts_pesq = artigos_por_pesquisa(pesquisa)
                dm_pesq = df_dm_com_preco[
                    df_dm_com_preco["artigo"].astype(str).isin(arts_pesq) |
                    df_dm_com_preco["artigo"].str.contains(pesquisa, case=False, na=False) |
                    df_dm_com_preco["familia"].str.contains(pesquisa, case=False, na=False) |
                    df_dm_com_preco["marca"].str.contains(pesquisa, case=False, na=False)
                ]
                marcas_cascade = sorted(dm_pesq["marca"].dropna().unique())
            else:
                marcas_cascade = marcas_disp
            marcas_sel = st.multiselect("Marca:", marcas_cascade, placeholder="Escolha Opção")
        with col_f:
            if pesquisa and not df_dm_com_preco.empty:
                if 'arts_pesq' not in locals():
                    arts_pesq = artigos_por_pesquisa(pesquisa)
                dm_base = df_dm_com_preco[
                    df_dm_com_preco["artigo"].astype(str).isin(arts_pesq) |
                    df_dm_com_preco["artigo"].str.contains(pesquisa, case=False, na=False) |
                    df_dm_com_preco["familia"].str.contains(pesquisa, case=False, na=False) |
                    df_dm_com_preco["marca"].str.contains(pesquisa, case=False, na=False)
                ].copy()
            else:
                dm_base = df_dm_com_preco.copy() if not df_dm_com_preco.empty else pd.DataFrame()
            if marcas_sel and not dm_base.empty:
                dm_base = dm_base[dm_base["marca"].isin(marcas_sel)]
            familias_cascade = sorted(dm_base["familia"].dropna().unique()) if not dm_base.empty else familias_disp
            familias_sel = st.multiselect("Família:", familias_cascade, placeholder="Escolha Opção")

        st.divider()

        col_pais, col_conc = st.columns([1, 4])
        with col_pais:
            st.markdown("**País:**")
            inc_pt = st.checkbox("Portugal", value=True, key="chk_pt")
            inc_es = st.checkbox("Espanha", value=True, key="chk_es")
            if conc_outros:
                inc_out = st.checkbox("Outros", value=True, key="chk_out")
            else:
                inc_out = False

        with col_conc:
            conc_por_pais = []
            if inc_pt:
                conc_por_pais += conc_pt
            if inc_es:
                conc_por_pais += conc_es
            if inc_out:
                conc_por_pais += conc_outros
            conc_sel = st.multiselect("Concorrentes:", conc_por_pais, default=conc_por_pais, format_func=formatar_nome_concorrente)

    if not conc_sel:
        st.info("Selecciona pelo menos um concorrente.")
    else:
        artigos_permitidos = None
        if (marcas_sel or familias_sel) and not df_dm.empty:
            mask = pd.Series([True] * len(df_dm), index=df_dm.index)
            if marcas_sel:
                mask &= df_dm["marca"].isin(marcas_sel)
            if familias_sel:
                mask &= df_dm["familia"].isin(familias_sel)
            artigos_permitidos = set(df_dm[mask]["artigo"].astype(str))

        with st.spinner("A carregar preços..."):
            df_raw = query_comparacao(tuple(conc_sel), pesquisa, tuple(marcas_sel), tuple(familias_sel))

        if artigos_permitidos is not None:
            df_raw = df_raw[df_raw["artigo"].astype(str).isin(artigos_permitidos)]

        if df_raw.empty:
            st.info("Nenhum artigo encontrado.")
        else:
            pivot_preco = df_raw.pivot_table(index=["artigo", "descricao"], columns="concorrente", values="preco", aggfunc="first").reset_index()
            pivot_preco.columns.name = None

            if not df_dm.empty:
                dm_sub = df_dm[["artigo", "pvp1", "marca", "familia"]].copy()
                dm_sub["artigo"] = dm_sub["artigo"].astype(str)
                pivot_preco["artigo"] = pivot_preco["artigo"].astype(str)
                pivot_preco = pivot_preco.merge(dm_sub, on="artigo", how="left")

            price_cols = [c for c in pivot_preco.columns if c in conc_sel]
            all_num_cols = price_cols + (["pvp1"] if "pvp1" in pivot_preco.columns else [])
            if all_num_cols:
                nums_all = pivot_preco[all_num_cols].apply(pd.to_numeric, errors="coerce")
                pivot_preco["_min_val"] = nums_all.min(axis=1)
                pivot_preco["_max_val"] = nums_all.max(axis=1)
                pivot_preco["Mínimo €"] = pivot_preco["_min_val"].round(2)

                def get_todos_mais_baratos(row):
                    mn = row["_min_val"]
                    if pd.isna(mn):
                        return "—"
                    res = []
                    if "pvp1" in row and pd.notna(row["pvp1"]) and abs(float(row["pvp1"]) - mn) < 0.005:
                        res.append("Douromed")
                    for c in price_cols:
                        try:
                            if pd.notna(row[c]) and abs(float(row[c]) - mn) < 0.005:
                                res.append(formatar_nome_concorrente(c))
                        except:
                            pass
                    return " · ".join(res) if res else "—"
                pivot_preco["Mais barato"] = pivot_preco.apply(get_todos_mais_baratos, axis=1)

            rename_cols = {c: formatar_nome_concorrente(c) for c in price_cols}
            rename_html = {c: flag_html(c) for c in price_cols}
            pivot_preco = pivot_preco.rename(columns=rename_cols)
            price_cols_fmt = [rename_cols[c] for c in price_cols]

            pivot_preco = pivot_preco.rename(columns={
                "artigo": "Artigo",
                "descricao": "Descrição",
                "pvp1": "PVP Dm",
                "marca": "Marca",
                "familia": "Família"
            })

            def fmt_preco(x):
                try:
                    if pd.isna(x) or x == "" or x == "—":
                        return "—"
                    return f"{float(x):.2f} €"
                except:
                    return "—"

            for col in price_cols_fmt + ["PVP Dm", "Mínimo €"]:
                if col in pivot_preco.columns:
                    pivot_preco[col] = pivot_preco[col].apply(fmt_preco)

            def extrair_num(s):
                try:
                    if pd.isna(s) or s == "—":
                        return None
                    return float(str(s).replace("€", "").replace(",", ".").strip())
                except:
                    return None

            cols_para_cor = price_cols_fmt + (["PVP Dm"] if "PVP Dm" in pivot_preco.columns else [])
            row_min_vals = {}
            row_max_vals = {}
            for idx2, row2 in pivot_preco.iterrows():
                vals = [extrair_num(row2.get(c)) for c in cols_para_cor]
                vals = [v for v in vals if v is not None]
                row_min_vals[idx2] = min(vals) if vals else None
                row_max_vals[idx2] = max(vals) if vals else None

            url_map = {}
            if not df_links.empty:
                for _, lr in df_links.iterrows():
                    url_map[(str(lr["artigo"]), lr["concorrente"])] = lr["url"]

            fmt_to_orig = {v: k for k, v in rename_cols.items()}
            cols_fixas = ["Artigo", "Descrição", "Marca", "Família", "PVP Dm", "Mínimo €", "Mais barato"]
            cols_final = [c for c in cols_fixas + price_cols_fmt if c in pivot_preco.columns]

            # CSS da tabela
            css_table = """
            <style>
            .dm-table-wrap { 
                overflow-x: auto !important; 
                max-height: 560px; 
                overflow-y: auto; 
                border: 1px solid #444; 
                border-radius: 8px; 
                width: 100%;
            }
            .dm-table { 
                border-collapse: collapse; 
                width: 100%; 
                min-width: 800px;
                font-family: 'Aptos', 'Segoe UI', 'Calibri', sans-serif; 
                font-size: 13px; 
                color: #e0e0e0; 
            }
            .dm-table thead th { 
                background: #1E3A5F; 
                color: #fff; 
                padding: 8px 12px; 
                white-space: nowrap; 
                position: sticky; 
                top: 0; 
                z-index: 2; 
            }
            .dm-table thead th.sortable { 
                cursor: pointer; 
                user-select: none; 
            }
            .dm-table thead th.sortable:hover { 
                background: #2a5298; 
            }
            .dm-table thead th .si { 
                margin-left: 4px; 
                font-size: 9px; 
                opacity: 0.6; 
            }
            .dm-table thead th.asc .si::after { 
                content: "▲"; 
                opacity: 1; 
            }
            .dm-table thead th.desc .si::after { 
                content: "▼"; 
                opacity: 1; 
            }
            .dm-table thead th.sortable:not(.asc):not(.desc) .si::after { 
                content: "⇅"; 
            }
            .dm-table thead th.left { 
                text-align: left; 
            }
            .dm-table thead th.center { 
                text-align: center; 
            }
            .dm-table tbody tr:nth-child(even) td { 
                background: #2a2a2a; 
            }
            .dm-table tbody tr:nth-child(odd) td { 
                background: #1e1e1e; 
            }
            .dm-table tbody tr:hover td { 
                background: #2e3d50 !important; 
            }
            .dm-table td { 
                padding: 6px 12px; 
                white-space: nowrap; 
                border-bottom: 1px solid #3a3a3a; 
            }
            .dm-table td.left { 
                text-align: left; 
            }
            .dm-table td.center { 
                text-align: center; 
            }
            a.dm-link { 
                color: #5b9bd5; 
                text-decoration: none; 
                font-weight: 500; 
            }
            a.dm-link:hover { 
                text-decoration: underline; 
            }
            .min-price { 
                color: #27ae60 !important; 
                font-weight: 700 !important; 
            }
            .max-price { 
                color: #e74c3c !important; 
                font-weight: 600 !important; 
            }
            </style>
            """

            rename_html_fmt = {rename_cols[c]: rename_html[c] for c in price_cols}
            SORTABLE = {"Artigo", "Descrição", "Marca", "Família"}

            th_rows = "<tr>"
            for ci, c in enumerate(cols_final):
                cls = "left" if c in ["Artigo", "Descrição", "Marca", "Família"] else "center"
                label = rename_html_fmt.get(c, c).replace("🦷 ", "")
                if c in SORTABLE:
                    th_rows += f"<th class='{cls} sortable' onclick='sort({ci},this)'>{label}<span class='si'></span></th>"
                else:
                    th_rows += f"<th class='{cls}'>{label}</th>"
            th_rows += "</tr>"

            linhas = []
            for idx2, row in pivot_preco.iterrows():
                artigo_val = str(row.get("Artigo", ""))
                row_min = row_min_vals.get(idx2)
                row_max = row_max_vals.get(idx2)
                tr = "<tr>"
                for c in cols_final:
                    val = row.get(c, "")
                    val_str = "—" if (val is None or pd.isna(val) or str(val).strip() in ("", "nan", "None")) else str(val)
                    cls = "left" if c in ["Artigo", "Descrição", "Marca", "Família"] else "center"
                    extra_class = ""
                    num_val = extrair_num(val_str) if c in cols_para_cor and val_str != "—" else None
                    if num_val is not None and row_min is not None:
                        if abs(num_val - row_min) < 0.005:
                            extra_class = "min-price"
                        elif row_max is not None and abs(num_val - row_max) < 0.005 and abs(row_max - row_min) > 0.005:
                            extra_class = "max-price"

                    if c in price_cols_fmt and val_str != "—":
                        orig = fmt_to_orig.get(c, c)
                        if orig in conc_sel:
                            url = url_map.get((artigo_val, orig), "")
                            if url and url.startswith("http"):
                                if extra_class == "min-price":
                                    cell = f"<a href='{url}' target='_blank' style='color:#27ae60;text-decoration:none;font-weight:700;'>{val_str}</a>"
                                elif extra_class == "max-price":
                                    cell = f"<a href='{url}' target='_blank' style='color:#e74c3c;text-decoration:none;font-weight:600;'>{val_str}</a>"
                                else:
                                    cell = f"<a href='{url}' target='_blank' class='dm-link'>{val_str}</a>"
                            else:
                                cell = val_str
                        else:
                            cell = val_str
                    elif c == "PVP Dm" and val_str != "—":
                        if extra_class == "min-price":
                            cell = f"<span style='color:#27ae60;font-weight:700;'>{val_str}</span>"
                        elif extra_class == "max-price":
                            cell = f"<span style='color:#e74c3c;font-weight:600;'>{val_str}</span>"
                        else:
                            cell = val_str
                    elif c == "Mínimo €" and val_str != "—":
                        cell = f"<span style='color:#27ae60;font-weight:700;'>{val_str}</span>"
                    elif c == "Mais barato" and val_str not in ("", "—", "nan"):
                        baratos = val_str.split(" · ")
                        links_html = []
                        for b in baratos:
                            if b == "Douromed":
                                links_html.append('<span style="color:#27ae60;font-weight:700;">Douromed</span>')
                            else:
                                raw_mb = b.replace("PT | ", "PT_").replace("ES | ", "ES_")
                                if raw_mb in conc_sel:
                                    url_mb = None
                                    for (art, conc), url in url_map.items():
                                        if art == artigo_val and (conc == raw_mb or formatar_nome_concorrente(conc) == b):
                                            url_mb = url
                                            break
                                    if url_mb and url_mb.startswith("http"):
                                        links_html.append(f"<a href='{url_mb}' target='_blank' style='color:#27ae60;text-decoration:none;font-weight:700;'>{b}</a>")
                                    else:
                                        links_html.append(f"<span style='color:#27ae60;font-weight:700;'>{b}</span>")
                                else:
                                    links_html.append(f"<span style='color:#27ae60;font-weight:700;'>{b}</span>")
                        cell = " · ".join(links_html) if links_html else "—"
                        extra_class = "min-price"
                    else:
                        cell = val_str

                    class_attr = f" class='{cls}'"
                    if extra_class == "min-price":
                        tr += f"<td{class_attr} style='color:#27ae60;font-weight:700;'>{cell}</td>"
                    elif extra_class == "max-price":
                        tr += f"<td{class_attr} style='color:#e74c3c;font-weight:600;'>{cell}</td>"
                    else:
                        tr += f"<td{class_attr}>{cell}</td>"
                tr += "</tr>"
                linhas.append(tr)
            rows_html = "".join(linhas)

            js_sort = """
            <script>
            function sort(idx, th) {
                var table = th.closest("table");
                var tbody = table.querySelector("tbody");
                var rows = Array.from(tbody.querySelectorAll("tr"));
                var asc = th.classList.contains("desc") || !th.classList.contains("asc");
                table.querySelectorAll("thead th").forEach(function(h){ h.classList.remove("asc","desc"); });
                th.classList.add(asc ? "asc" : "desc");
                rows.sort(function(a,b){
                    var va = (a.cells[idx] ? a.cells[idx].innerText : "").trim();
                    var vb = (b.cells[idx] ? b.cells[idx].innerText : "").trim();
                    return asc ? va.localeCompare(vb,"pt") : vb.localeCompare(va,"pt");
                });
                rows.forEach(function(r){ tbody.appendChild(r); });
            }
            </script>
            """
            n_rows = len(pivot_preco)
            altura = min(600, 80 + n_rows * 35)
            tabela_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">{css_table}{js_sort}</head><body style="margin:0;padding:0;background:transparent;"><div class="dm-table-wrap" style="max-height:{altura}px;"><table class="dm-table" id="tbl"><thead>{th_rows}</thead><tbody>{rows_html}</tbody></table></div></body></html>"""
            components.html(tabela_html, height=altura + 20, scrolling=False)
            st.caption(f"{n_rows} artigos · {ultima_data.strftime('%d/%m/%Y %H:%M') if ultima_data != 'N/A' else 'Sem dados'}")

            if st.button("📥 Exportar para Excel"):
                from io import BytesIO
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                from openpyxl.utils import get_column_letter

                buf = BytesIO()

                preco_map = {}
                ref_bd_map = {}
                if not df_raw.empty:
                    for _, row in df_raw.iterrows():
                        key = (str(row["artigo"]), row["concorrente"])
                        preco_map[key] = row["preco"]
                        ref = row.get("referencia", "")
                        if pd.isna(ref) or str(ref).lower() in ("nan", "none", "n/a"):
                            ref_bd_map[key] = ""
                        else:
                            ref_bd_map[key] = str(ref).strip()

                cols_fixas_export = ["Artigo", "Descrição", "Marca", "Família", "PVP Dm", "Mínimo €", "Mais barato"]
                df_export = pivot_preco[cols_fixas_export].copy()

                colunas_preco = []
                colunas_ref = []
                for conc_original in conc_sel:
                    col_preco_fmt = formatar_nome_concorrente(conc_original)
                    if col_preco_fmt in pivot_preco.columns:
                        col_ref = f"{col_preco_fmt} (Ref)"
                        df_export[col_preco_fmt] = df_export.apply(
                            lambda row, c=conc_original: preco_map.get((str(row["Artigo"]), c), None), axis=1
                        )
                        colunas_preco.append(col_preco_fmt)
                        df_export[col_ref] = df_export.apply(
                            lambda row, c=conc_original: ref_bd_map.get((str(row["Artigo"]), c), ""), axis=1
                        )
                        colunas_ref.append(col_ref)

                df_export = df_export.fillna("")

                wb = Workbook()
                ws = wb.active
                ws.title = "Comparação"

                for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        if isinstance(value, str) and value.strip().upper() == "N/A":
                            value = ""
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '0.00'

                header_row = [cell.value for cell in ws[1]]
                for col_ref in colunas_ref:
                    if col_ref in header_row:
                        col_idx = header_row.index(col_ref) + 1
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                cell.number_format = '@'

                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(buf)
                buf.seek(0)

                st.download_button(
                    "Descarregar",
                    buf.getvalue(),
                    file_name=f"comparacao_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# ============================================================
# TAB 2 – EVOLUÇÃO DE PREÇOS (COM FILTROS)
# ============================================================
with tab2:
    if not todos_art:
        st.info("Nenhum artigo encontrado.")
    else:
        dict_info_dm_tab2 = {}
        if not df_dm.empty:
            for _, row in df_dm.iterrows():
                dict_info_dm_tab2[row["artigo"]] = {
                    "descricao": row.get("descricao", ""),
                    "marca": row.get("marca", ""),
                    "familia": row.get("familia", "")
                }

        if "tab2_inc_pt" not in st.session_state:
            st.session_state.tab2_inc_pt = True
        if "tab2_inc_es" not in st.session_state:
            st.session_state.tab2_inc_es = True
        if "tab2_inc_out" not in st.session_state:
            st.session_state.tab2_inc_out = False
        if "tab2_aplicar" not in st.session_state:
            st.session_state.tab2_aplicar = 0
        if "tab2_filtro_conc" not in st.session_state:
            iniciais = []
            if st.session_state.tab2_inc_pt:
                iniciais.extend(conc_pt)
            if st.session_state.tab2_inc_es:
                iniciais.extend(conc_es)
            if st.session_state.tab2_inc_out:
                iniciais.extend(conc_outros)
            iniciais.sort(key=lambda x: (0 if x.startswith("PT_") else 1 if x.startswith("ES_") else 2, x))
            st.session_state.tab2_filtro_conc = iniciais.copy()
        if "tab2_artigo_sel" not in st.session_state:
            st.session_state.tab2_artigo_sel = None

        def ordem_pais(nome):
            if nome.startswith("PT_"): return (0, nome)
            elif nome.startswith("ES_"): return (1, nome)
            else: return (2, nome)

        st.subheader("📈 Evolução de Preços")

        with st.expander("🔎 Filtros", expanded=True):
            col_periodo, col_pesquisa = st.columns([1, 2])
            with col_periodo:
                periodo = st.selectbox("📅 Período", ["7 dias", "15 dias", "30 dias", "60 dias", "90 dias", "Todo o histórico"], index=4, key="tab2_periodo")
            with col_pesquisa:
                pesquisa = st.text_input("🔍 Artigo / Descrição:", key="tab2_pesquisa", placeholder="Digite o código ou descrição...")

            st.divider()

            col_f2, col_f3 = st.columns(2)
            with col_f2:
                marcas_sel = st.multiselect("🏷️ Marca:", marcas_disp, key="tab2_marcas", placeholder="Escolha Opção")
            with col_f3:
                familias_sel = st.multiselect("📁 Família:", familias_disp, key="tab2_familias", placeholder="Escolha Opção")

            st.divider()

            col_pais, col_conc = st.columns([1, 4])
            with col_pais:
                st.markdown("**🌍 País:**")
                inc_pt = st.checkbox("Portugal", value=st.session_state.tab2_inc_pt, key="tab2_chk_pt")
                inc_es = st.checkbox("Espanha", value=st.session_state.tab2_inc_es, key="tab2_chk_es")
                inc_out = False
                if conc_outros:
                    inc_out = st.checkbox("Outros", value=st.session_state.tab2_inc_out, key="tab2_chk_out")

                if (inc_pt != st.session_state.tab2_inc_pt or inc_es != st.session_state.tab2_inc_es or inc_out != st.session_state.tab2_inc_out):
                    st.session_state.tab2_inc_pt = inc_pt
                    st.session_state.tab2_inc_es = inc_es
                    st.session_state.tab2_inc_out = inc_out

                    nova_selecao = set(st.session_state.tab2_filtro_conc)
                    if inc_pt: nova_selecao.update(conc_pt)
                    else: nova_selecao.difference_update(conc_pt)
                    if inc_es: nova_selecao.update(conc_es)
                    else: nova_selecao.difference_update(conc_es)
                    if inc_out: nova_selecao.update(conc_outros)
                    else: nova_selecao.difference_update(conc_outros)

                    st.session_state.tab2_filtro_conc = sorted(list(nova_selecao), key=ordem_pais)
                    st.session_state.tab2_aplicar += 1

            with col_conc:
                opcoes_conc = []
                if inc_pt: opcoes_conc.extend(conc_pt)
                if inc_es: opcoes_conc.extend(conc_es)
                if inc_out: opcoes_conc.extend(conc_outros)
                opcoes_conc = sorted(opcoes_conc, key=ordem_pais)

                selecao_valida = [c for c in st.session_state.tab2_filtro_conc if c in opcoes_conc]
                if set(selecao_valida) != set(st.session_state.tab2_filtro_conc):
                    st.session_state.tab2_filtro_conc = selecao_valida

                multiselect_key = f"tab2_conc_{st.session_state.tab2_aplicar}"
                conc_sel = st.multiselect("🏢 Concorrentes:", options=opcoes_conc, default=st.session_state.tab2_filtro_conc,
                                          format_func=formatar_nome_concorrente, key=multiselect_key, placeholder="Escolha Opção")
                st.session_state.tab2_filtro_conc = conc_sel

        if not st.session_state.tab2_filtro_conc:
            st.warning("⚠️ Selecione pelo menos um concorrente para visualizar a evolução de preços.")
        else:
            artigos_filtrados = set(todos_art)
            if pesquisa:
                pesquisa_lower = pesquisa.lower()
                artigos_por_pesquisa = set()
                for art in todos_art:
                    info = dict_info_dm_tab2.get(art, {})
                    desc = info.get("descricao", "").lower()
                    if pesquisa_lower in art.lower() or pesquisa_lower in desc:
                        artigos_por_pesquisa.add(art)
                artigos_filtrados &= artigos_por_pesquisa
            if marcas_sel:
                artigos_por_marca = set()
                for art in todos_art:
                    if dict_info_dm_tab2.get(art, {}).get("marca") in marcas_sel:
                        artigos_por_marca.add(art)
                artigos_filtrados &= artigos_por_marca
            if familias_sel:
                artigos_por_familia = set()
                for art in todos_art:
                    if dict_info_dm_tab2.get(art, {}).get("familia") in familias_sel:
                        artigos_por_familia.add(art)
                artigos_filtrados &= artigos_por_familia

            lista_artigos_final = sorted(artigos_filtrados)

            if not lista_artigos_final:
                st.warning("Nenhum artigo corresponde aos filtros selecionados.")
            else:
                if st.session_state.tab2_artigo_sel not in lista_artigos_final:
                    st.session_state.tab2_artigo_sel = lista_artigos_final[0]

                artigo_sel = st.selectbox("📦 Artigo", lista_artigos_final,
                                         index=lista_artigos_final.index(st.session_state.tab2_artigo_sel),
                                         format_func=lambda x: f"{x} - {dict_info_dm_tab2.get(x, {}).get('descricao', '')[:60]}",
                                         key="tab2_artigo_select")
                st.session_state.tab2_artigo_sel = artigo_sel

                dias_map = {"7 dias": 7, "15 dias": 15, "30 dias": 30, "60 dias": 60, "90 dias": 90, "Todo o histórico": 9999}
                with st.spinner("A carregar histórico..."):
                    df_hist = query_historico(artigo_sel, tuple(st.session_state.tab2_filtro_conc), dias_map[periodo])

                if df_hist.empty:
                    st.info("Sem histórico para este artigo no período selecionado.")
                else:
                    info_artigo = dict_info_dm_tab2.get(artigo_sel, {})
                    df_hist["artigo"] = artigo_sel
                    df_hist["descricao"] = info_artigo.get("descricao", "")
                    df_hist["marca"] = info_artigo.get("marca", "")
                    df_hist["familia"] = info_artigo.get("familia", "")

                    pvp_dm = None
                    if not df_dm.empty:
                        row_dm = df_dm[df_dm["artigo"] == artigo_sel]
                        if not row_dm.empty:
                            pvp_dm = row_dm["pvp1"].values[0]

                    df_hist["concorrente_fmt"] = df_hist["concorrente"].apply(formatar_nome_concorrente)
                    fig = px.line(df_hist, x="data", y="preco", color="concorrente_fmt", markers=True,
                                  labels={"data": "Data", "preco": "Preço (€)", "concorrente_fmt": "Concorrente"},
                                  title=f"Evolução de preços — {artigo_sel}")
                    if pvp_dm and not pd.isna(pvp_dm):
                        fig.add_hline(y=pvp_dm, line_dash="dash", line_color="green",
                                      annotation_text=f"PVP Douromed: {pvp_dm:.2f}€", annotation_position="right")
                    fig.update_layout(legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                                      yaxis_tickformat=".2f", hovermode="x unified")
                    st.plotly_chart(fig, use_container_width=True)

                    with st.expander("📋 Histórico de Preços", expanded=False):
                        df_tab = df_hist.sort_values(["concorrente", "data"])
                        df_tab["preco_anterior"] = df_tab.groupby("concorrente")["preco"].shift(1)
                        html_linhas = ""
                        for _, row in df_tab.sort_values("data", ascending=False).iterrows():
                            data_str = row["data"].strftime("%d/%m/%Y %H:%M")
                            data_sort = row["data"].strftime("%Y%m%d%H%M")
                            conc_str = formatar_nome_concorrente(row["concorrente"])
                            preco_val = row["preco"] if pd.notna(row["preco"]) else None
                            preco_str = f"{preco_val:.2f} €" if preco_val is not None else "—"
                            preco_ant = row["preco_anterior"]
                            variacao_num = None
                            if pd.isna(preco_ant):
                                variacao_html = '<span style="color: #888;">—</span>'
                            elif row["preco"] > preco_ant:
                                dif = row["preco"] - preco_ant
                                pct = (dif / preco_ant) * 100
                                variacao_num = pct
                                variacao_html = f'<span style="color: #e74c3c; font-weight: bold;">▲ {preco_ant:.2f}€ → {row["preco"]:.2f}€ (+{pct:.1f}%)</span>'
                            elif row["preco"] < preco_ant:
                                dif = preco_ant - row["preco"]
                                pct = (dif / preco_ant) * 100
                                variacao_num = -pct
                                variacao_html = f'<span style="color: #27ae60; font-weight: bold;">▼ {preco_ant:.2f}€ → {row["preco"]:.2f}€ (-{pct:.1f}%)</span>'
                            else:
                                variacao_num = 0
                                variacao_html = f'<span style="color: #f39c12; font-weight: bold;">= {row["preco"]:.2f}€ → {row["preco"]:.2f}€ (0.0%)</span>'

                            html_linhas += f"""
                            <tr data-date="{data_sort}" data-preco="{preco_val if preco_val else 0}" data-variacao="{variacao_num if variacao_num is not None else ''}">
                                <td>{data_str}</td>
                                <td>{conc_str}</td>
                                <td>{preco_str}</td>
                                <td>{variacao_html}</td>
                            </tr>
                            """
                        html_historico = f"""
                        <!DOCTYPE html><html><head><meta charset="UTF-8"><style>
                            .tabela-container {{ max-height: 400px; overflow-y: auto; border: 1px solid #444; border-radius: 8px; background-color: #1e1e1e; }}
                            table {{ width: 100%; border-collapse: collapse; font-family: 'Aptos', 'Segoe UI', 'Calibri', sans-serif; font-size: 13px; background-color: #1e1e1e; color: #e0e0e0; }}
                            th {{ background-color: #1E3A5F; color: white; padding: 10px 12px; text-align: center; font-weight: bold; position: sticky; top: 0; z-index: 10; cursor: pointer; user-select: none; }}
                            th:hover {{ background-color: #2a5298; }}
                            th span {{ margin-left: 5px; font-size: 10px; opacity: 0.7; }}
                            td {{ padding: 8px 12px; text-align: center; border-bottom: 1px solid #444; }}
                            tr:nth-child(even) {{ background-color: #2a2a2a; }}
                            tr:nth-child(odd) {{ background-color: #1e1e1e; }}
                            tr:hover {{ background-color: #2e3d50 !important; }}
                            .asc::after {{ content: " ▲"; opacity: 1; }}
                            .desc::after {{ content: " ▼"; opacity: 1; }}
                        </style>
                        <script>
                        function sortTable(tableId, columnIndex, type) {{
                            var table = document.getElementById(tableId);
                            var tbody = table.querySelector("tbody");
                            var rows = Array.from(tbody.querySelectorAll("tr"));
                            var header = table.querySelectorAll("th")[columnIndex];
                            var isAsc = header.classList.contains("asc");
                            table.querySelectorAll("th").forEach(th => th.classList.remove("asc", "desc"));
                            if (isAsc) {{
                                header.classList.add("desc");
                            }} else {{
                                header.classList.add("asc");
                            }}
                            rows.sort(function(a, b) {{
                                var aVal, bVal;
                                if (type === 'date') {{
                                    aVal = a.getAttribute("data-date") || "";
                                    bVal = b.getAttribute("data-date") || "";
                                    return isAsc ? bVal.localeCompare(aVal) : aVal.localeCompare(bVal);
                                }} else if (type === 'number') {{
                                    aVal = parseFloat(a.getAttribute("data-preco")) || 0;
                                    bVal = parseFloat(b.getAttribute("data-preco")) || 0;
                                    return isAsc ? bVal - aVal : aVal - bVal;
                                }} else if (type === 'variacao') {{
                                    aVal = parseFloat(a.getAttribute("data-variacao")) || -Infinity;
                                    bVal = parseFloat(b.getAttribute("data-variacao")) || -Infinity;
                                    if (isNaN(aVal)) aVal = isAsc ? Infinity : -Infinity;
                                    if (isNaN(bVal)) bVal = isAsc ? Infinity : -Infinity;
                                    return isAsc ? bVal - aVal : aVal - bVal;
                                }} else {{
                                    aVal = a.cells[columnIndex].innerText.trim();
                                    bVal = b.cells[columnIndex].innerText.trim();
                                    return isAsc ? bVal.localeCompare(aVal, 'pt', {{numeric: true}}) : aVal.localeCompare(bVal, 'pt', {{numeric: true}});
                                }}
                            }});
                            rows.forEach(row => tbody.appendChild(row));
                        }}
                        </script>
                        </head><body style="margin:0;padding:0;background-color:#1e1e1e;">
                            <div class="tabela-container">
                                <table id="historicoTable">
                                    <thead>
                                        <tr>
                                            <th onclick="sortTable('historicoTable', 0, 'date')">Data <span>⇅</span></th>
                                            <th onclick="sortTable('historicoTable', 1, 'text')">Concorrente <span>⇅</span></th>
                                            <th onclick="sortTable('historicoTable', 2, 'number')">Preço <span>⇅</span></th>
                                            <th onclick="sortTable('historicoTable', 3, 'variacao')">Variação <span>⇅</span></th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {html_linhas}
                                    </tbody>
                                </table>
                            </div>
                        </body></html>
                        """
                        components.html(html_historico, height=450, scrolling=False)
                        st.caption(f"Total: {len(df_tab)} registos")

                    with st.expander("📊 Resumo por Concorrente", expanded=False):
                        grupos = df_hist.groupby("concorrente")
                        resumo_data = []
                        for conc, grupo in grupos:
                            idx_min = grupo["preco"].idxmin() if not grupo["preco"].isna().all() else None
                            idx_max = grupo["preco"].idxmax() if not grupo["preco"].isna().all() else None
                            min_val = grupo.loc[idx_min, "preco"] if idx_min is not None else None
                            max_val = grupo.loc[idx_max, "preco"] if idx_max is not None else None
                            mean_val = grupo["preco"].mean()
                            count_val = len(grupo)
                            data_min = grupo.loc[idx_min, "data"] if idx_min is not None else None
                            data_max = grupo.loc[idx_max, "data"] if idx_max is not None else None
                            resumo_data.append({
                                "concorrente": conc,
                                "Mínimo": min_val,
                                "Máximo": max_val,
                                "Média": mean_val,
                                "Registos": count_val,
                                "data_min": data_min,
                                "data_max": data_max
                            })

                        stats = pd.DataFrame(resumo_data)
                        stats["concorrente_fmt"] = stats["concorrente"].apply(formatar_nome_concorrente)
                        stats = stats.sort_values("concorrente_fmt")

                        html_resumo_rows = ""
                        for _, row in stats.iterrows():
                            conc_fmt = row["concorrente_fmt"]
                            minimo = row["Mínimo"]
                            maximo = row["Máximo"]
                            media = row["Média"]
                            registos = int(row["Registos"])
                            data_min = row["data_min"]
                            data_max = row["data_max"]

                            if pd.notna(minimo):
                                data_min_str = pd.to_datetime(data_min).strftime("%d/%m/%Y") if pd.notna(data_min) else ""
                                minimo_str = f"{minimo:.2f} €"
                                if data_min_str:
                                    minimo_str += f" <span style='font-size:0.8em; color:#aaa;'>({data_min_str})</span>"
                            else:
                                minimo_str = "—"

                            if pd.notna(maximo):
                                data_max_str = pd.to_datetime(data_max).strftime("%d/%m/%Y") if pd.notna(data_max) else ""
                                maximo_str = f"{maximo:.2f} €"
                                if data_max_str:
                                    maximo_str += f" <span style='font-size:0.8em; color:#aaa;'>({data_max_str})</span>"
                            else:
                                maximo_str = "—"

                            media_str = f"{media:.2f} €" if pd.notna(media) else "—"

                            min_num = minimo if pd.notna(minimo) else 0
                            max_num = maximo if pd.notna(maximo) else 0
                            mean_num = media if pd.notna(media) else 0

                            minimo_td = f'<td style="color: #27ae60; font-weight: 600;">{minimo_str}</td>' if minimo_str != "—" else f'<td>{minimo_str}</td>'
                            maximo_td = f'<td style="color: #e74c3c; font-weight: 600;">{maximo_str}</td>' if maximo_str != "—" else f'<td>{maximo_str}</td>'

                            html_resumo_rows += f"""
                            <tr data-min="{min_num}" data-max="{max_num}" data-mean="{mean_num}" data-count="{registos}">
                                <td style="text-align: left;">{conc_fmt}</td>
                                {minimo_td}
                                {maximo_td}
                                <td>{media_str}</td>
                                <td>{registos}</td>
                            </tr>
                            """

                        html_resumo = f"""
                        <!DOCTYPE html><html><head><meta charset="UTF-8"><style>
                            .resumo-container {{ max-height: 400px; overflow-y: auto; border: 1px solid #444; border-radius: 8px; background-color: #1e1e1e; }}
                            table {{ width: 100%; border-collapse: collapse; font-family: 'Aptos', 'Segoe UI', 'Calibri', sans-serif; font-size: 13px; background-color: #1e1e1e; color: #e0e0e0; }}
                            th {{ background-color: #1E3A5F; color: white; padding: 8px 10px; text-align: center; font-weight: bold; position: sticky; top: 0; z-index: 10; cursor: pointer; user-select: none; }}
                            th:hover {{ background-color: #2a5298; }}
                            th span {{ margin-left: 5px; font-size: 10px; opacity: 0.7; }}
                            td {{ padding: 6px 10px; text-align: center; border-bottom: 1px solid #444; }}
                            tr:nth-child(even) {{ background-color: #2a2a2a; }}
                            tr:nth-child(odd) {{ background-color: #1e1e1e; }}
                            tr:hover {{ background-color: #2e3d50 !important; }}
                            .asc::after {{ content: " ▲"; opacity: 1; }}
                            .desc::after {{ content: " ▼"; opacity: 1; }}
                        </style>
                        <script>
                        function sortResumo(columnIndex, type) {{
                            var table = document.getElementById("resumoTable");
                            var tbody = table.querySelector("tbody");
                            var rows = Array.from(tbody.querySelectorAll("tr"));
                            var header = table.querySelectorAll("th")[columnIndex];
                            var isAsc = header.classList.contains("asc");
                            table.querySelectorAll("th").forEach(th => th.classList.remove("asc", "desc"));
                            if (isAsc) {{
                                header.classList.add("desc");
                            }} else {{
                                header.classList.add("asc");
                            }}
                            rows.sort(function(a, b) {{
                                var aVal, bVal;
                                if (type === 'text') {{
                                    aVal = a.cells[columnIndex].innerText.trim();
                                    bVal = b.cells[columnIndex].innerText.trim();
                                    return isAsc ? bVal.localeCompare(aVal, 'pt') : aVal.localeCompare(bVal, 'pt');
                                }} else if (type === 'min') {{
                                    aVal = parseFloat(a.getAttribute("data-min")) || 0;
                                    bVal = parseFloat(b.getAttribute("data-min")) || 0;
                                    return isAsc ? bVal - aVal : aVal - bVal;
                                }} else if (type === 'max') {{
                                    aVal = parseFloat(a.getAttribute("data-max")) || 0;
                                    bVal = parseFloat(b.getAttribute("data-max")) || 0;
                                    return isAsc ? bVal - aVal : aVal - bVal;
                                }} else if (type === 'mean') {{
                                    aVal = parseFloat(a.getAttribute("data-mean")) || 0;
                                    bVal = parseFloat(b.getAttribute("data-mean")) || 0;
                                    return isAsc ? bVal - aVal : aVal - bVal;
                                }} else if (type === 'count') {{
                                    aVal = parseInt(a.getAttribute("data-count")) || 0;
                                    bVal = parseInt(b.getAttribute("data-count")) || 0;
                                    return isAsc ? bVal - aVal : aVal - bVal;
                                }}
                            }});
                            rows.forEach(row => tbody.appendChild(row));
                        }}
                        </script>
                        </head><body style="margin:0;padding:0;background-color:#1e1e1e;">
                            <div class="resumo-container">
                                <table id="resumoTable">
                                    <thead>
                                        <tr>
                                            <th onclick="sortResumo(0, 'text')">Concorrente <span>⇅</span></th>
                                            <th onclick="sortResumo(1, 'min')">Mínimo <span>⇅</span></th>
                                            <th onclick="sortResumo(2, 'max')">Máximo <span>⇅</span></th>
                                            <th onclick="sortResumo(3, 'mean')">Média <span>⇅</span></th>
                                            <th onclick="sortResumo(4, 'count')">Registos <span>⇅</span></th>
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {html_resumo_rows}
                                    </tbody>
                                </table>
                            </div>
                        </body></html>
                        """
                        components.html(html_resumo, height=200 + len(stats)*35, scrolling=True)

# ============================================================
# TAB 3 – ALERTAS
# ============================================================
with tab3:
    st.subheader("📊 Variações de preço (última vs penúltima recolha)")

    df_alertas = query_alertas()

    if df_alertas.empty:
        st.info("Sem variações de preço registadas.")
    else:
        df_alertas["data_atual"] = pd.to_datetime(df_alertas["data_atual"])
        df_alertas["data_anterior"] = pd.to_datetime(df_alertas["data_anterior"])
        df_alertas["data_atual_str"] = df_alertas["data_atual"].dt.strftime("%Y-%m-%d")
        df_alertas["hora_atual_str"] = df_alertas["data_atual"].dt.strftime("%H:%M")
        df_alertas["data_anterior_str"] = df_alertas["data_anterior"].dt.strftime("%Y-%m-%d")
        df_alertas["hora_anterior_str"] = df_alertas["data_anterior"].dt.strftime("%H:%M")

        # Usar o dicionário do Douromed para obter marca/família
        dict_douromed = {}
        if not df_dm.empty:
            df_dm["artigo"] = df_dm["artigo"].astype(str).str.strip()
            for _, row in df_dm.iterrows():
                artigo = row["artigo"]
                dict_douromed[artigo] = {
                    "descricao": row.get("descricao", ""),
                    "marca": row.get("marca", ""),
                    "familia": row.get("familia", "")
                }

        marcas_lista = []
        familias_lista = []
        for _, row in df_alertas.iterrows():
            artigo = row['artigo']
            info_dm = dict_douromed.get(artigo, {})
            marcas_lista.append(info_dm.get("marca", ""))
            familias_lista.append(info_dm.get("familia", ""))

        df_alertas["marca"] = marcas_lista
        df_alertas["familia"] = familias_lista

        todos_concorrentes = sorted(df_alertas["concorrente"].unique())
        conc_pt = [c for c in todos_concorrentes if c.upper().startswith("PT_")]
        conc_es = [c for c in todos_concorrentes if c.upper().startswith("ES_")]
        conc_outros = [c for c in todos_concorrentes if not c.upper().startswith(("PT_", "ES_"))]

        if "tab3_inc_pt" not in st.session_state:
            st.session_state.tab3_inc_pt = True
        if "tab3_inc_es" not in st.session_state:
            st.session_state.tab3_inc_es = True
        if "tab3_inc_out" not in st.session_state:
            st.session_state.tab3_inc_out = False
        if "tab3_aplicar" not in st.session_state:
            st.session_state.tab3_aplicar = 0
        if "tab3_pesquisa_anterior" not in st.session_state:
            st.session_state.tab3_pesquisa_anterior = ""

        with st.expander("🔎 Filtros", expanded=True):
            col_periodo, col_pesquisa = st.columns([1, 2])
            with col_periodo:
                periodo = st.selectbox("📅 Período:", ["7 dias", "15 dias", "30 dias", "60 dias", "90 dias", "Todo o histórico"], index=4, key="tab3_periodo")
            with col_pesquisa:
                pesquisa = st.text_input("🔍 Artigo / Descrição:", key="tab3_pesquisa", placeholder="Digite o código ou descrição...")
                if pesquisa != st.session_state.tab3_pesquisa_anterior:
                    st.session_state.tab3_pesquisa_anterior = pesquisa
                    st.session_state.tab3_aplicar += 1

            st.divider()

            col_f2, col_f3 = st.columns(2)
            df_filtrado_temp = df_alertas.copy()
            if periodo != "Todo o histórico":
                dias_map = {"7 dias": 7, "15 dias": 15, "30 dias": 30, "60 dias": 60, "90 dias": 90}
                dias = dias_map.get(periodo, 90)
                data_limite = datetime.now() - timedelta(days=dias)
                df_filtrado_temp = df_filtrado_temp[df_filtrado_temp["data_atual"] >= data_limite]

            if pesquisa:
                pesquisa_lower = pesquisa.lower()
                df_filtrado_temp = df_filtrado_temp[
                    df_filtrado_temp["artigo"].str.lower().str.contains(pesquisa_lower, na=False) |
                    df_filtrado_temp["descricao"].str.lower().str.contains(pesquisa_lower, na=False)
                ]

            marcas_disponiveis = sorted([m for m in df_filtrado_temp["marca"].dropna().unique() if m and m != "nan"])
            with col_f2:
                marcas_sel = st.multiselect("🏷️ Marca:", marcas_disponiveis, key="tab3_marcas", placeholder="Escolha Opção")
            if marcas_sel:
                df_filtrado_temp = df_filtrado_temp[df_filtrado_temp["marca"].isin(marcas_sel)]

            familias_disponiveis = sorted([f for f in df_filtrado_temp["familia"].dropna().unique() if f and f != "nan"])
            with col_f3:
                familias_sel = st.multiselect("📁 Família:", familias_disponiveis, key="tab3_familias", placeholder="Escolha Opção")
            if familias_sel:
                df_filtrado_temp = df_filtrado_temp[df_filtrado_temp["familia"].isin(familias_sel)]

            st.divider()

            col_pais, col_conc = st.columns([1, 4])
            with col_pais:
                st.markdown("**🌍 País:**")
                inc_pt = st.checkbox("Portugal", value=st.session_state.tab3_inc_pt, key="tab3_chk_pt")
                inc_es = st.checkbox("Espanha", value=st.session_state.tab3_inc_es, key="tab3_chk_es")
                if conc_outros:
                    inc_out = st.checkbox("Outros", value=st.session_state.tab3_inc_out, key="tab3_chk_out")
                else:
                    inc_out = False

                pt_mudou = inc_pt != st.session_state.tab3_inc_pt
                es_mudou = inc_es != st.session_state.tab3_inc_es
                out_mudou = inc_out != st.session_state.tab3_inc_out
                st.session_state.tab3_inc_pt = inc_pt
                st.session_state.tab3_inc_es = inc_es
                st.session_state.tab3_inc_out = inc_out
                if pt_mudou or es_mudou or out_mudou:
                    st.session_state.tab3_aplicar += 1

            with col_conc:
                conc_por_pais = []
                conc_pt_filtrados = [c for c in conc_pt if c in df_filtrado_temp["concorrente"].unique()]
                conc_es_filtrados = [c for c in conc_es if c in df_filtrado_temp["concorrente"].unique()]
                conc_outros_filtrados = [c for c in conc_outros if c in df_filtrado_temp["concorrente"].unique()]

                if inc_pt: conc_por_pais += conc_pt_filtrados
                if inc_es: conc_por_pais += conc_es_filtrados
                if inc_out: conc_por_pais += conc_outros_filtrados

                multiselect_key = f"tab3_conc_{st.session_state.tab3_aplicar}"
                default_concs = conc_por_pais.copy()
                conc_sel = st.multiselect("🏢 Concorrentes:", conc_por_pais, default=default_concs,
                                          format_func=formatar_nome_concorrente, key=multiselect_key, placeholder="Escolha Opção")

        # Aplicar filtros ao DataFrame final
        df_filtrado = df_alertas.copy()
        if periodo != "Todo o histórico":
            dias_map = {"7 dias": 7, "15 dias": 15, "30 dias": 30, "60 dias": 60, "90 dias": 90}
            dias = dias_map.get(periodo, 90)
            data_limite = datetime.now() - timedelta(days=dias)
            df_filtrado = df_filtrado[df_filtrado["data_atual"] >= data_limite]

        if pesquisa:
            pesquisa_lower = pesquisa.lower()
            df_filtrado = df_filtrado[
                df_filtrado["artigo"].str.lower().str.contains(pesquisa_lower, na=False) |
                df_filtrado["descricao"].str.lower().str.contains(pesquisa_lower, na=False)
            ]
        if marcas_sel:
            df_filtrado = df_filtrado[df_filtrado["marca"].isin(marcas_sel)]
        if familias_sel:
            df_filtrado = df_filtrado[df_filtrado["familia"].isin(familias_sel)]
        if conc_sel:
            df_filtrado = df_filtrado[df_filtrado["concorrente"].isin(conc_sel)]

        df_filtrado = df_filtrado.sort_values("data_atual", ascending=False)

        if df_filtrado.empty:
            st.info("Nenhuma variação encontrada com os filtros selecionados.")
        else:
            descidas = df_filtrado[df_filtrado["variacao_pct"] < 0].sort_values("data_atual", ascending=False)
            subidas = df_filtrado[df_filtrado["variacao_pct"] > 0].sort_values("data_atual", ascending=False)

            st.caption(f"📊 {len(df_filtrado)} variações encontradas | 📉 {len(descidas)} descidas | 📈 {len(subidas)} subidas")
            st.markdown("---")

            col_d, col_s = st.columns(2)

            with col_d:
                st.markdown("#### 📉 Descidas de Preço")
                if descidas.empty:
                    st.info("Sem descidas com os filtros selecionados.")
                else:
                    for _, r in descidas.iterrows():
                        variacao = r['variacao_pct']
                        cor = "#27ae60"
                        seta = "▼"

                        artigo_display = f"**{r['artigo']}**"
                        if r.get('descricao') and r['descricao'] not in ("", "nan"):
                            artigo_display = f"**{r['artigo']}** - {r['descricao'][:80]}"

                        info_extra = ""
                        if r.get('marca') and r['marca'] not in ("", "nan"):
                            info_extra += f"🏷️ {r['marca']} "
                        if r.get('familia') and r['familia'] not in ("", "nan"):
                            info_extra += f"📁 {r['familia']}"

                        data_hora_atual = f"{r['data_atual_str']} {r['hora_atual_str']}"
                        data_hora_anterior = f"{r['data_anterior_str']} {r['hora_anterior_str']}"

                        conc_nome = formatar_nome_concorrente(r['concorrente'])
                        url_produto = None
                        if not df_links.empty:
                            mask = (df_links['artigo'].astype(str) == str(r['artigo'])) & (df_links['concorrente'] == r['concorrente'])
                            urls = df_links.loc[mask, 'url']
                            if not urls.empty:
                                url_produto = urls.iloc[0]

                        if url_produto and isinstance(url_produto, str) and url_produto.startswith('http'):
                            conc_html = f"<a href='{url_produto}' target='_blank' style='color:#666; font-size:0.9em; text-decoration:none;'>{conc_nome}</a>"
                        else:
                            conc_html = f"<span style='color:#666; font-size:0.9em;'>{conc_nome}</span>"

                        st.markdown(
                            f"""{artigo_display}<br>
                            {conc_html}<br>
                            <span style='color:#888; font-size:0.8em;'>{info_extra}</span><br>
                            <span style='color:#666; font-size:0.85em;'>📅 {data_hora_anterior} → {data_hora_atual}</span><br>
                            <span style='color:{cor}; font-weight:bold;'>{seta} {r['preco_anterior']:.2f}€ → {r['preco_atual']:.2f}€ ({variacao:+.1f}%)</span>
                            """,
                            unsafe_allow_html=True
                        )
                        st.markdown("---")

            with col_s:
                st.markdown("#### 📈 Subidas de Preço")
                if subidas.empty:
                    st.info("Sem subidas com os filtros selecionados.")
                else:
                    for _, r in subidas.iterrows():
                        variacao = r['variacao_pct']
                        cor = "#e74c3c"
                        seta = "▲"

                        artigo_display = f"**{r['artigo']}**"
                        if r.get('descricao') and r['descricao'] not in ("", "nan"):
                            artigo_display = f"**{r['artigo']}** - {r['descricao'][:80]}"

                        info_extra = ""
                        if r.get('marca') and r['marca'] not in ("", "nan"):
                            info_extra += f"🏷️ {r['marca']} "
                        if r.get('familia') and r['familia'] not in ("", "nan"):
                            info_extra += f"📁 {r['familia']}"

                        data_hora_atual = f"{r['data_atual_str']} {r['hora_atual_str']}"
                        data_hora_anterior = f"{r['data_anterior_str']} {r['hora_anterior_str']}"

                        conc_nome = formatar_nome_concorrente(r['concorrente'])
                        url_produto = None
                        if not df_links.empty:
                            mask = (df_links['artigo'].astype(str) == str(r['artigo'])) & (df_links['concorrente'] == r['concorrente'])
                            urls = df_links.loc[mask, 'url']
                            if not urls.empty:
                                url_produto = urls.iloc[0]

                        if url_produto and isinstance(url_produto, str) and url_produto.startswith('http'):
                            conc_html = f"<a href='{url_produto}' target='_blank' style='color:#666; font-size:0.9em; text-decoration:none;'>{conc_nome}</a>"
                        else:
                            conc_html = f"<span style='color:#666; font-size:0.9em;'>{conc_nome}</span>"

                        st.markdown(
                            f"""{artigo_display}<br>
                            {conc_html}<br>
                            <span style='color:#888; font-size:0.8em;'>{info_extra}</span><br>
                            <span style='color:#666; font-size:0.85em;'>📅 {data_hora_anterior} → {data_hora_atual}</span><br>
                            <span style='color:{cor}; font-weight:bold;'>{seta} {r['preco_anterior']:.2f}€ → {r['preco_atual']:.2f}€ ({variacao:+.1f}%)</span>
                            """,
                            unsafe_allow_html=True
                        )
                        st.markdown("---")

# ============================================================
# TAB 4 – GESTÃO DE CONCORRENTES
# ============================================================
with tab4:
    st.subheader("⚙️ Gestão de Concorrentes")

    concorrentes_ativos = listar_concorrentes_ativos()

    # SEÇÃO 1: ATUALIZAR CATÁLOGO (chama script externo adaptado)
    with st.expander("🔄 Actualizar Catálogo (Douromed + Concorrentes)", expanded=False):
        st.markdown("""
        Esta operação irá:
        - 🔄 Actualizar a tabela **artigos** com os dados mais recentes do SQL Server da Douromed.
        - 📝 Actualizar as tabelas **concorrentes** e **links** com base nas folhas do Excel.
        """)
        st.info("⚠️ Esta operação pode demorar alguns minutos. O script 'gerar_template_excel.py' deve estar adaptado para escrever na BD do Neon.")

        if st.button("🔄 Actualizar Catálogo", type="primary", key="btn_actualizar_catalogo"):
            with st.spinner("A actualizar catálogo..."):
                try:
                    script_path = BASE_DIR / "gerar_template_excel.py"
                    if script_path.exists():
                        result = subprocess.run(
                            [sys.executable, str(script_path)],
                            capture_output=True,
                            text=True,
                            timeout=300,
                            encoding='utf-8',
                            errors='replace'
                        )
                        log_output = result.stdout + "\n" + result.stderr if result.stderr else result.stdout
                        st.text_area("Log da actualização:", log_output, height=200)
                        if result.returncode == 0:
                            st.success("✅ Catálogo actualizado com sucesso!")
                            st.cache_data.clear()
                            st.rerun()
                        else:
                            st.error(f"❌ Erro ao actualizar catálogo. Código de saída: {result.returncode}")
                    else:
                        st.error(f"❌ Script não encontrado em: {script_path}")
                except Exception as e:
                    st.error(f"❌ Erro: {str(e)}")

    st.divider()

    # SEÇÃO 2: ELIMINAR CONCORRENTE
    with st.expander("❌ Eliminar Concorrente", expanded=False):
        lista_eliminar = sorted([c[0] for c in concorrentes_ativos])
        if not lista_eliminar:
            st.info("Nenhum concorrente ativo para eliminar.")
        else:
            opcoes_select = [""] + lista_eliminar
            concorrente_eliminar = st.selectbox(
                "Seleccione o concorrente a eliminar:",
                options=opcoes_select,
                format_func=lambda x: "Escolha Opção" if x == "" else formatar_nome_concorrente(x),
                key="eliminar_select"
            )
            if "confirmar_eliminar" not in st.session_state:
                st.session_state.confirmar_eliminar = False

            if not st.session_state.confirmar_eliminar:
                btn_disabled = (concorrente_eliminar == "")
                if st.button("🗑️ Eliminar Concorrente", type="primary", key="btn_eliminar", disabled=btn_disabled):
                    st.session_state.confirmar_eliminar = True
                    st.rerun()
            else:
                st.warning(f"⚠️ Tem a certeza que deseja eliminar permanentemente '{formatar_nome_concorrente(concorrente_eliminar)}'?")
                col_sim, col_nao = st.columns(2)
                with col_sim:
                    if st.button("✅ Sim, eliminar", type="primary", key="btn_confirmar_eliminar"):
                        with st.spinner(f"A eliminar {concorrente_eliminar}..."):
                            sucesso, registos, _, erro_msg = eliminar_concorrente_bd(concorrente_eliminar)
                            if sucesso:
                                st.success(f"✅ Concorrente '{formatar_nome_concorrente(concorrente_eliminar)}' eliminado!")
                                if registos > 0:
                                    st.info(f"📊 {registos} concorrentes removidos da BD (preços e links associados também foram apagados).")
                                st.session_state.confirmar_eliminar = False
                                st.cache_data.clear()
                                st.rerun()
                            else:
                                st.error(f"❌ Erro: {erro_msg or 'Falha desconhecida.'}")
                with col_nao:
                    if st.button("❌ Cancelar", key="btn_cancelar_eliminar"):
                        st.session_state.confirmar_eliminar = False
                        st.rerun()

    st.divider()

    # SEÇÃO 3: ADICIONAR NOVO CONCORRENTE
    with st.expander("➕ Adicionar Novo Concorrente", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            novo_concorrente = st.text_input("Nome do concorrente (ex: PT_NovoConcorrente):", placeholder="PT_Nome ou ES_Nome", key="input_novo_concorrente")
        with col2:
            url_homepage = st.text_input("🔗 URL da homepage:", placeholder="https://www.exemplo.com", key="input_nova_homepage")
        st.caption("O nome deve começar com PT_ ou ES_.")
        if st.button("➕ Adicionar", type="primary", key="btn_adicionar_concorrente"):
            if not novo_concorrente:
                st.error("❌ O nome do concorrente é obrigatório.")
            elif not novo_concorrente.startswith(("PT_", "ES_")):
                st.error("❌ O nome deve começar com PT_ ou ES_")
            elif not url_homepage or not url_homepage.startswith("http"):
                st.error("❌ A URL da homepage é obrigatória e deve começar com http:// ou https://")
            elif novo_concorrente in [c[0] for c in concorrentes_ativos]:
                st.error("❌ Esse concorrente já existe na lista de ativos.")
            else:
                with st.spinner(f"A adicionar {novo_concorrente}..."):
                    sucesso, mensagem = adicionar_concorrente_bd(novo_concorrente, url_homepage)
                    if sucesso:
                        st.success(f"✅ {mensagem}")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(f"❌ {mensagem}")

    st.divider()

    # SEÇÃO 4: LISTA DE CONCORRENTES ATIVOS
    with st.expander("📋 Concorrentes Ativos", expanded=False):
        if concorrentes_ativos:
            conc_pt_lista = [c for c in concorrentes_ativos if c[0].startswith("PT_")]
            conc_es_lista = [c for c in concorrentes_ativos if c[0].startswith("ES_")]
            col_pt, col_es = st.columns(2)
            with col_pt:
                st.markdown("**Portugal**")
                if conc_pt_lista:
                    for nome, url in sorted(conc_pt_lista):
                        nome_fmt = formatar_nome_concorrente(nome)
                        if url:
                            st.markdown(f"- [{nome_fmt}]({url})", unsafe_allow_html=True)
                        else:
                            st.markdown(f"- {nome_fmt}")
                else:
                    st.write("*Nenhum*")
            with col_es:
                st.markdown("**Espanha**")
                if conc_es_lista:
                    for nome, url in sorted(conc_es_lista):
                        nome_fmt = formatar_nome_concorrente(nome)
                        if url:
                            st.markdown(f"- [{nome_fmt}]({url})", unsafe_allow_html=True)
                        else:
                            st.markdown(f"- {nome_fmt}")
                else:
                    st.write("*Nenhum*")
            st.caption(f"Total: {len(concorrentes_ativos)} concorrentes ativos")
        else:
            st.info("Nenhum concorrente ativo configurado.")

# ============================================================
# TAB 5 – GESTÃO DE LINKS E REFERÊNCIAS
# ============================================================
with tab5:
    st.subheader("🔗 Gestão de Links e Referências por Artigo")
    st.caption("Adicione ou edite o URL e a referência manual do produto para cada concorrente.")

    if df_dm.empty:
        st.warning("Não foi possível carregar os artigos da Douromed.")
    else:
        dict_info_dm = {}
        for _, row in df_dm.iterrows():
            dict_info_dm[row["artigo"]] = {
                "descricao": row.get("descricao", ""),
                "marca": row.get("marca", ""),
                "familia": row.get("familia", "")
            }
        artigos_lista = sorted(dict_info_dm.keys())

        # Estado da sessão
        if 'conc_sel_gestao' not in st.session_state:
            st.session_state.conc_sel_gestao = None
        if 'artigo_sel_gestao' not in st.session_state:
            st.session_state.artigo_sel_gestao = None

        with st.expander("🔎 Filtros", expanded=True):
            col1, col2 = st.columns([2, 2])
            with col1:
                conc_list = [None] + [c[0] for c in listar_concorrentes_ativos()]
                idx_conc = 0
                if st.session_state.conc_sel_gestao in conc_list:
                    idx_conc = conc_list.index(st.session_state.conc_sel_gestao)
                st.selectbox("🏢 Concorrente", conc_list,
                             format_func=lambda x: formatar_nome_concorrente(x) if x else "Selecione...",
                             index=idx_conc, key="conc_widget")
                st.session_state.conc_sel_gestao = st.session_state.conc_widget
            with col2:
                pesquisa = st.text_input("🔍 Pesquisar artigo", placeholder="Código ou descrição...", key="pesquisa_input")

        if not st.session_state.conc_sel_gestao:
            st.info("🏢 Selecione um concorrente para começar.")
        else:
            conc_atual = st.session_state.conc_sel_gestao

            arts_filtrados = set(artigos_lista)
            if pesquisa:
                pesquisa_lower = pesquisa.lower()
                arts_filtrados = {a for a in arts_filtrados if pesquisa_lower in a.lower() or
                                  pesquisa_lower in dict_info_dm.get(a, {}).get("descricao", "").lower()}
            arts_filtrados = sorted(arts_filtrados)

            if not arts_filtrados:
                st.info("Nenhum artigo encontrado.")
            else:
                if st.session_state.artigo_sel_gestao not in arts_filtrados:
                    st.session_state.artigo_sel_gestao = arts_filtrados[0]
                artigo_sel = st.selectbox("📦 Artigo", arts_filtrados,
                                         index=arts_filtrados.index(st.session_state.artigo_sel_gestao),
                                         format_func=lambda x: f"{x} - {dict_info_dm.get(x, {}).get('descricao', '')[:60]}",
                                         key="art_widget")
                st.session_state.artigo_sel_gestao = artigo_sel

                # Obter dados atuais da BD
                engine = obter_conn()
                with engine.connect() as conn:
                    result = conn.execute(text("SELECT url, referencia_manual FROM links WHERE artigo = :art AND concorrente = :conc"),
                                          {"art": artigo_sel, "conc": conc_atual})
                    row = result.fetchone()
                url_atual = row[0] if row else ""
                ref_manual_atual = row[1] if row else ""

                info_dm = dict_info_dm.get(artigo_sel, {})
                @st.cache_data(ttl=3600)
                def obter_ultimo_preco(artigo: str, concorrente: str):
                    eng = obter_conn()
                    with eng.connect() as conn:
                        res = conn.execute(text("""
                            SELECT preco, data, referencia FROM precos
                            WHERE sucesso=1 AND artigo = :art AND concorrente = :conc
                            ORDER BY data DESC LIMIT 1
                        """), {"art": artigo, "conc": concorrente})
                        r = res.fetchone()
                    if r and r[0] and float(r[0]) > 0:
                        preco_val = float(r[0])
                        data_raw = r[1] if r[1] else ""
                        referencia = r[2] if len(r) > 2 and r[2] else ""
                        return f"{preco_val:.2f} € ({data_raw})", preco_val, data_raw, referencia
                    return "Sem registo de preço", None, None, ""
                ultimo_preco, _, _, _ = obter_ultimo_preco(artigo_sel, conc_atual)

                st.markdown("---")
                col_info, col_edit = st.columns([1, 2])
                with col_info:
                    st.markdown(f"**📝 Descrição:** {info_dm.get('descricao', '')}")
                    st.markdown(f"**💰 Último Preço:** {ultimo_preco}")
                    if info_dm.get("marca"): st.markdown(f"**🏷️ Marca:** {info_dm['marca']}")
                    if info_dm.get("familia"): st.markdown(f"**📁 Família:** {info_dm['familia']}")
                    st.markdown(f"**🏢 Concorrente:** {formatar_nome_concorrente(conc_atual)}")

                with col_edit:
                    novo_url = st.text_input("🔗 URL do produto", value=url_atual, placeholder="https://...", key="url_input")
                    nova_ref = st.text_input("🏷️ Referência (manual)", value=ref_manual_atual, placeholder="Ex: 16-144", key="ref_input")
                    col_btn1, col_btn2, col_btn3 = st.columns(3)
                    with col_btn1:
                        if st.button("💾 Guardar", use_container_width=True, key="save_link"):
                            if not novo_url or not novo_url.startswith("http"):
                                st.error("❌ URL inválido")
                            else:
                                try:
                                    with engine.connect() as conn:
                                        conn.execute(text("""
                                            INSERT INTO links (artigo, concorrente, url, referencia_manual)
                                            VALUES (:art, :conc, :url, :ref)
                                            ON CONFLICT (artigo, concorrente)
                                            DO UPDATE SET url = :url, referencia_manual = :ref
                                        """), {"art": artigo_sel, "conc": conc_atual, "url": novo_url, "ref": nova_ref})
                                        conn.commit()
                                    st.success("✅ Link e referência guardados!")
                                    st.cache_data.clear()
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Erro: {e}")
                    with col_btn2:
                        if st.button("🗑️ Limpar", use_container_width=True, key="clear_link"):
                            try:
                                with engine.connect() as conn:
                                    conn.execute(text("UPDATE links SET url = '', referencia_manual = '' WHERE artigo = :art AND concorrente = :conc"),
                                                 {"art": artigo_sel, "conc": conc_atual})
                                    conn.commit()
                                st.success("✅ URL e referência removidos!")
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Erro: {e}")
                    with col_btn3:
                        if url_atual and url_atual.startswith("http"):
                            st.markdown(f"[🔗 Abrir]({url_atual})", unsafe_allow_html=True)

# ============================================================
# TAB 6 – ADMINISTRAÇÃO
# ============================================================
with tab6:
    st.subheader("🛠️ Administração do Sistema")
    st.caption("Ferramentas para gestão avançada do Dental Market.")

    if "admin_autenticado" not in st.session_state:
        st.session_state.admin_autenticado = False

    if not st.session_state.admin_autenticado:
        senha = st.text_input("🔐 Introduza a senha de administrador:", type="password")
        if st.button("Entrar"):
            if senha == ADMIN_PASSWORD:
                st.session_state.admin_autenticado = True
                st.success("Acesso concedido.")
                st.rerun()
            else:
                st.error("Senha incorreta.")
    else:
        if st.sidebar.button("🚪 Sair da Administração"):
            st.session_state.admin_autenticado = False
            st.rerun()

        # Executar scraper manual
        with st.expander("🕷️ Executar Scraper", expanded=True):
            st.markdown("### Execução manual do scraper")
            conc_list = ["Todos"] + [c[0] for c in listar_concorrentes_ativos()]
            conc_selecionado = st.selectbox("Concorrente:", conc_list,
                                           format_func=lambda x: "Todos os concorrentes" if x == "Todos" else formatar_nome_concorrente(x))
            teste_qtd = st.number_input("Limitar a N produtos (0 = todos)", min_value=0, value=0, step=1)
            if st.button("🚀 Executar Scraper Agora", type="primary"):
                if conc_selecionado == "Todos":
                    site_arg = ""
                else:
                    site_arg = conc_selecionado
                comando = [sys.executable, "scraper.py"]
                if site_arg:
                    comando.extend(["--site", site_arg])
                if teste_qtd > 0:
                    comando.extend(["--teste", str(teste_qtd)])
                env = os.environ.copy()
                env["PYTHONIOENCODING"] = "utf-8"
                with st.spinner(f"A executar scraper para {conc_selecionado}..."):
                    try:
                        resultado = subprocess.run(comando, cwd=BASE_DIR, capture_output=True, timeout=3600,
                                                   env=env, encoding='utf-8', errors='replace')
                        st.success("Execução concluída!")
                        st.text_area("Log de saída:", resultado.stdout, height=300)
                        if resultado.stderr:
                            st.text_area("Erros:", resultado.stderr, height=150)
                        st.cache_data.clear()
                        st.rerun()
                    except subprocess.TimeoutExpired:
                        st.error("Timeout: o scraper demorou mais de 1 hora.")
                    except Exception as e:
                        st.error(f"Erro ao executar scraper: {e}")

            st.markdown("---")
            st.markdown("### Histórico de Execuções")
            if SCRAPER_LOG.exists():
                with open(SCRAPER_LOG, "r", encoding="utf-8") as f:
                    linhas = f.readlines()
                if linhas:
                    linhas_recentes = linhas[-10:][::-1]
                    log_text = "".join(linhas_recentes)
                    st.text_area("Últimas execuções:", log_text, height=200, disabled=True)
                else:
                    st.info("Nenhuma execução registada.")
            else:
                st.info("Nenhum registo encontrado.")

        st.divider()

        # Limpar base de dados
        with st.expander("🗑️ Limpar Base de Dados", expanded=False):
            st.warning("⚠️ Esta operação apagará **todos os registos de preços** da base de dados.")
            if "confirmar_limpar_bd" not in st.session_state:
                st.session_state.confirmar_limpar_bd = False

            if not st.session_state.confirmar_limpar_bd:
                if st.button("🗑️ Limpar Base de Dados", type="secondary"):
                    st.session_state.confirmar_limpar_bd = True
                    st.rerun()
            else:
                st.error("**Tem a certeza absoluta?**")
                col_sim, col_nao = st.columns(2)
                with col_sim:
                    if st.button("✅ Sim, apagar tudo", type="primary"):
                        with st.spinner("A apagar registos..."):
                            try:
                                engine = obter_conn()
                                with engine.connect() as conn:
                                    conn.execute(text("DELETE FROM precos"))
                                    conn.commit()
                                st.session_state.confirmar_limpar_bd = False
                                st.cache_data.clear()
                                st.success("✅ Registos apagados com sucesso!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Erro ao limpar base de dados: {e}")
                with col_nao:
                    if st.button("❌ Cancelar"):
                        st.session_state.confirmar_limpar_bd = False
                        st.rerun()

        st.divider()

        with st.expander("ℹ️ Informações do Sistema", expanded=False):
            st.markdown(f"**Base de dados:** PostgreSQL (Neon)")
            st.markdown(f"**Python:** `{sys.executable}`")
            st.markdown(f"**Streamlit:** `{st.__version__}`")